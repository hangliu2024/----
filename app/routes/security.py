from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from app.decorators import department_permission_required, admin_required
from app.models import (PermissionMatrix, ClassifiedPersonnel, ClassifiedMedia, 
                        SecurityZone, ElectronicDocument, PaperDocument,
                        SysRole, SysModule, SysPermission, SysUserRole, SysDataPermission,
                        PersonSystemPermissionMatrix)
from app.decorators import clear_permission_cache
from app import db
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

bp = Blueprint('security', __name__)

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def safe_int(value, default=None):
    """安全转换为整数，用于dept_id等字段"""
    if value is None or value == '':
        return default
    try:
        return int(value)
    except (ValueError, TypeError):
        return default

@bp.route('/security')
@login_required
@admin_required
def index():
    # 统计数据
    permission_count = SysPermission.query.count()
    role_count = SysRole.query.filter_by(status=1).count()
    module_count = SysModule.query.filter_by(status=1).count()
    personnel_count = ClassifiedPersonnel.query.count()
    media_count = ClassifiedMedia.query.count()
    zone_count = SecurityZone.query.count()
    electronic_count = ElectronicDocument.query.count()
    paper_count = PaperDocument.query.count()
    
    # 获取角色列表
    roles = SysRole.query.filter_by(status=1).order_by(SysRole.sort_order).all()
    
    # 获取模块树
    modules = SysModule.query.filter_by(status=1).order_by(SysModule.sort_order).all()
    
    return render_template('security/index.html',
                           permission_count=permission_count,
                           role_count=role_count,
                           module_count=module_count,
                           personnel_count=personnel_count,
                           media_count=media_count,
                           zone_count=zone_count,
                           electronic_count=electronic_count,
                           paper_count=paper_count,
                           roles=roles,
                           modules=modules)

# ==================== 角色管理 ====================

@bp.route('/security/roles')
@login_required
@admin_required
def role_list():
    """角色列表"""
    roles = SysRole.query.order_by(SysRole.sort_order).all()
    return render_template('security/role_list.html', roles=roles)

@bp.route('/security/roles/add', methods=['GET', 'POST'])
@login_required
@admin_required
def role_add():
    """添加角色"""
    if request.method == 'POST':
        role = SysRole(
            role_name=request.form.get('role_name'),
            role_code=request.form.get('role_code'),
            description=request.form.get('description'),
            status=int(request.form.get('status', 1)),
            sort_order=int(request.form.get('sort_order', 0)),
            created_by=current_user.id
        )
        db.session.add(role)
        db.session.commit()
        flash('角色添加成功！', 'success')
        return redirect(url_for('security.role_list'))
    return render_template('security/role_form.html')

@bp.route('/security/roles/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def role_edit(id):
    """编辑角色"""
    role = SysRole.query.get_or_404(id)
    if request.method == 'POST':
        role.role_name = request.form.get('role_name')
        role.description = request.form.get('description')
        role.status = int(request.form.get('status', 1))
        role.sort_order = int(request.form.get('sort_order', 0))
        db.session.commit()
        flash('角色更新成功！', 'success')
        return redirect(url_for('security.role_list'))
    return render_template('security/role_form.html', role=role)

@bp.route('/security/roles/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def role_delete(id):
    """删除角色"""
    role = SysRole.query.get_or_404(id)
    if role.role_code == 'super_admin':
        flash('超级管理员角色不能删除！', 'danger')
        return redirect(url_for('security.role_list'))
    
    # 检查是否有用户分配了该角色
    user_roles = SysUserRole.query.filter_by(role_id=id).all()
    if user_roles:
        affected_user_ids = [ur.user_id for ur in user_roles]
        # 先删除关联关系
        SysUserRole.query.filter_by(role_id=id).delete()
        SysDataPermission.query.filter_by(role_id=id).delete()
        # 清除受影响用户的权限缓存
        for uid in affected_user_ids:
            clear_permission_cache(uid)
    
    db.session.delete(role)
    db.session.commit()
    flash('角色删除成功！', 'success')
    return redirect(url_for('security.role_list'))

# ==================== 模块管理 ====================

@bp.route('/security/modules')
@login_required
@admin_required
def module_list():
    """模块列表"""
    modules = SysModule.query.order_by(SysModule.parent_id, SysModule.sort_order).all()
    return render_template('security/module_list.html', modules=modules)

@bp.route('/security/modules/add', methods=['GET', 'POST'])
@login_required
@admin_required
def module_add():
    """添加模块"""
    if request.method == 'POST':
        module = SysModule(
            module_name=request.form.get('module_name'),
            module_code=request.form.get('module_code'),
            parent_id=int(request.form.get('parent_id', 0)),
            module_type=request.form.get('module_type', 'menu'),
            route_path=request.form.get('route_path'),
            icon=request.form.get('icon'),
            sort_order=int(request.form.get('sort_order', 0)),
            status=int(request.form.get('status', 1))
        )
        db.session.add(module)
        db.session.commit()
        flash('模块添加成功！', 'success')
        return redirect(url_for('security.module_list'))
    
    # 获取父模块选项
    parent_modules = SysModule.query.filter_by(parent_id=0, status=1).all()
    return render_template('security/module_form.html', parent_modules=parent_modules)

@bp.route('/security/modules/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def module_edit(id):
    """编辑模块"""
    module = SysModule.query.get_or_404(id)
    if request.method == 'POST':
        module.module_name = request.form.get('module_name')
        module.parent_id = int(request.form.get('parent_id', 0))
        module.module_type = request.form.get('module_type', 'menu')
        module.route_path = request.form.get('route_path')
        module.icon = request.form.get('icon')
        module.sort_order = int(request.form.get('sort_order', 0))
        module.status = int(request.form.get('status', 1))
        db.session.commit()
        flash('模块更新成功！', 'success')
        return redirect(url_for('security.module_list'))
    
    parent_modules = SysModule.query.filter_by(parent_id=0, status=1).all()
    return render_template('security/module_form.html', module=module, parent_modules=parent_modules)

@bp.route('/security/permission-matrix')
@login_required
@admin_required
def role_permission_config():
    """角色权限矩阵配置页面 - 可视化配置每个角色对各模块的增删改查权限"""

    roles = SysRole.query.filter_by(status=1).order_by(SysRole.sort_order).all()
    modules = SysModule.query.filter_by(status=1).order_by(SysModule.parent_id, SysModule.sort_order).all()

    all_perms = SysPermission.query.all()
    perm_map = {}
    for p in all_perms:
        key = f'{p.role_id}_{p.module_id}'
        perm_map[key] = {
            'view': p.can_view, 'add': p.can_add,
            'edit': p.can_edit, 'delete': p.can_delete,
            'export': p.can_export, 'import': p.can_import,
            'audit': p.can_audit, 'approve': p.can_approve
        }

    return render_template('security/role_permission_config.html',
                           roles=roles, modules=modules, perm_map=perm_map)


@bp.route('/security/permissions/batch_save', methods=['POST'])
@login_required
@admin_required
def permissions_batch_save():
    """批量保存角色权限配置"""

    data = request.get_json()
    items = data.get('permissions', [])

    for item in items:
        role_id = item['role_id']
        module_id = item['module_id']
        actions = item['actions']

        perm = SysPermission.query.filter_by(
            role_id=role_id, module_id=module_id
        ).first()

        if not perm:
            perm = SysPermission(role_id=role_id, module_id=module_id)
            db.session.add(perm)

        if 'view' in actions: perm.can_view = actions['view']
        if 'add' in actions: perm.can_add = actions['add']
        if 'edit' in actions: perm.can_edit = actions['edit']
        if 'delete' in actions: perm.can_delete = actions['delete']
        if 'export' in actions: perm.can_export = actions['export']
        if 'import' in actions: perm.can_import = actions['import']
        if 'audit' in actions: perm.can_audit = actions['audit']
        if 'approve' in actions: perm.can_approve = actions['approve']

    db.session.commit()
    return jsonify({'success': True, 'message': f'已保存 {len(items)} 条权限配置'})


# ==================== 权限配置 ====================

@bp.route('/security/permissions')
@login_required
@admin_required
def permission_list():
    """人员权限矩阵 - 员工对公司各业务系统的权限配置"""
    search = request.args.get('search', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    employees_query = db.session.query(
        PersonSystemPermissionMatrix.emp_id,
        PersonSystemPermissionMatrix.emp_name,
        PersonSystemPermissionMatrix.dept_id,
        PersonSystemPermissionMatrix.dept_name
    ).distinct()
    
    if search:
        employees_query = employees_query.filter(
            db.or_(
                PersonSystemPermissionMatrix.emp_id.contains(search),
                PersonSystemPermissionMatrix.emp_name.contains(search),
                PersonSystemPermissionMatrix.dept_name.contains(search)
            )
        )
    
    employees_query = employees_query.order_by(PersonSystemPermissionMatrix.emp_id)
    
    total_emp = employees_query.count()
    employees_pagination = employees_query.offset((page - 1) * per_page).limit(per_page).all()
    
    systems = db.session.query(PersonSystemPermissionMatrix.system_name).distinct().all()
    systems = [s[0] for s in systems]
    
    emp_ids = [emp[0] for emp in employees_pagination]
    
    if emp_ids:
        permissions = PersonSystemPermissionMatrix.query.filter(
            PersonSystemPermissionMatrix.emp_id.in_(emp_ids)
        ).all()
    else:
        permissions = []
    
    permission_dict = {}
    for perm in permissions:
        key = f'["{perm.emp_id}", "{perm.system_name}"]'
        permission_dict[key] = {
            'can_view': perm.can_view,
            'can_add': perm.can_add,
            'can_edit': perm.can_edit,
            'can_delete': perm.can_delete,
            'can_export': perm.can_export,
            'can_import': perm.can_import,
            'can_approve': perm.can_approve,
            'can_config': perm.can_config
        }
    
    total_pages = (total_emp + per_page - 1) // per_page
    
    return render_template('security/person_permission_matrix.html',
                           employees=employees_pagination,
                           systems=systems,
                           permission_dict=permission_dict,
                           permission_count=len(permissions),
                           search=search,
                           page=page,
                           total_pages=total_pages,
                           total_emp=total_emp,
                           per_page=per_page)

@bp.route('/security/person_permission/edit/<emp_id>')
@login_required
@admin_required
def person_permission_edit(emp_id):
    """编辑单个员工的权限"""
    emp_perms = PersonSystemPermissionMatrix.query.filter_by(emp_id=emp_id).all()
    
    if emp_perms:
        emp = emp_perms[0]
        emp_name = emp.emp_name
        dept_id = emp.dept_id
        dept_name = emp.dept_name
    else:
        emp_name = ''
        dept_id = ''
        dept_name = ''
    
    all_systems = db.session.query(PersonSystemPermissionMatrix.system_name).distinct().all()
    all_systems = [s[0] for s in all_systems]
    
    perm_dict = {}
    for perm in emp_perms:
        perm_dict[perm.system_name] = perm
    
    return render_template('security/person_permission_edit.html',
                           emp_id=emp_id,
                           emp_name=emp_name,
                           dept_id=dept_id,
                           dept_name=dept_name,
                           all_systems=all_systems,
                           perm_dict=perm_dict)

@bp.route('/security/person_permission/update', methods=['POST'])
@login_required
@admin_required
def person_permission_update():
    """更新人员权限配置"""
    perm_id = request.form.get('id')
    perm = PersonSystemPermissionMatrix.query.get(perm_id)
    
    if not perm:
        return jsonify({'success': False, 'message': '权限配置不存在'})
    
    perm.can_view = 1 if request.form.get('can_view') else 0
    perm.can_add = 1 if request.form.get('can_add') else 0
    perm.can_edit = 1 if request.form.get('can_edit') else 0
    perm.can_delete = 1 if request.form.get('can_delete') else 0
    perm.can_export = 1 if request.form.get('can_export') else 0
    perm.can_import = 1 if request.form.get('can_import') else 0
    perm.can_approve = 1 if request.form.get('can_approve') else 0
    perm.can_config = 1 if request.form.get('can_config') else 0
    perm.permission_level = request.form.get('permission_level', 'basic')
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '权限配置更新成功'})

@bp.route('/security/person_permission/save', methods=['POST'])
@login_required
@admin_required
def person_permission_save():
    """保存员工的所有权限（批量优化版）"""
    emp_id = request.form.get('emp_id')
    emp_name = request.form.get('emp_name')
    dept_id = request.form.get('dept_id') or ''
    dept_name = request.form.get('dept_name') or ''

    systems = db.session.query(PersonSystemPermissionMatrix.system_name).distinct().all()
    systems = [s[0] for s in systems]

    existing_map = {}
    existing_records = PersonSystemPermissionMatrix.query.filter_by(emp_id=emp_id).all()
    for r in existing_records:
        existing_map[r.system_name] = r

    for system in systems:
        can_view = 1 if request.form.get('can_view_' + system) else 0
        can_add = 1 if request.form.get('can_add_' + system) else 0
        can_edit = 1 if request.form.get('can_edit_' + system) else 0
        can_delete = 1 if request.form.get('can_delete_' + system) else 0
        can_export = 1 if request.form.get('can_export_' + system) else 0
        can_import = 1 if request.form.get('can_import_' + system) else 0
        can_approve = 1 if request.form.get('can_approve_' + system) else 0
        can_config = 1 if request.form.get('can_config_' + system) else 0

        existing = existing_map.get(system)
        if existing:
            existing.emp_name = emp_name
            existing.dept_id = dept_id
            existing.dept_name = dept_name
            existing.can_view = can_view
            existing.can_add = can_add
            existing.can_edit = can_edit
            existing.can_delete = can_delete
            existing.can_export = can_export
            existing.can_import = can_import
            existing.can_approve = can_approve
            existing.can_config = can_config
        else:
            if can_view or can_add or can_edit or can_delete or can_export or can_import or can_approve or can_config:
                perm = PersonSystemPermissionMatrix(
                    emp_id=emp_id, emp_name=emp_name,
                    dept_id=dept_id, dept_name=dept_name,
                    system_name=system,
                    can_view=can_view, can_add=can_add,
                    can_edit=can_edit, can_delete=can_delete,
                    can_export=can_export, can_import=can_import,
                    can_approve=can_approve, can_config=can_config,
                    permission_level='basic'
                )
                db.session.add(perm)

    db.session.commit()
    return jsonify({'success': True, 'message': '权限保存成功'})

@bp.route('/security/person_permission/delete', methods=['POST'])
@login_required
@admin_required
def person_permission_delete():
    """删除人员权限配置"""
    perm_id = request.form.get('id')
    perm = PersonSystemPermissionMatrix.query.get(perm_id)
    
    if not perm:
        return jsonify({'success': False, 'message': '权限配置不存在'})
    
    db.session.delete(perm)
    db.session.commit()
    
    return jsonify({'success': True, 'message': '权限配置删除成功'})

@bp.route('/security/person_permission/list')
@login_required
@admin_required
def person_permission_list():
    """获取权限列表（支持按部门筛选）"""
    dept_filter = request.args.get('dept_id')
    system_filter = request.args.get('system_name')
    
    query = PersonSystemPermissionMatrix.query
    
    if dept_filter:
        query = query.filter_by(dept_id=dept_filter)
    if system_filter:
        query = query.filter_by(system_name=system_filter)
    
    permissions = query.all()
    
    return jsonify({
        'success': True,
        'data': [{
            'id': p.id,
            'emp_id': p.emp_id,
            'emp_name': p.emp_name,
            'dept_id': p.dept_id,
            'dept_name': p.dept_name,
            'system_name': p.system_name,
            'can_view': p.can_view,
            'can_add': p.can_add,
            'can_edit': p.can_edit,
            'can_delete': p.can_delete,
            'can_export': p.can_export,
            'can_import': p.can_import,
            'can_approve': p.can_approve,
            'can_config': p.can_config,
            'permission_level': p.permission_level
        } for p in permissions]
    })

@bp.route('/security/person_permission/batch_update', methods=['POST'])
@login_required
@admin_required
def person_permission_batch_update():
    """批量更新员工权限"""
    emp_id = request.form.get('emp_id')
    system_name = request.form.get('system_name')
    permissions = request.form.getlist('permissions')
    
    perm = PersonSystemPermissionMatrix.query.filter_by(
        emp_id=emp_id, system_name=system_name
    ).first()
    
    if not perm:
        perm = PersonSystemPermissionMatrix(
            emp_id=emp_id,
            system_name=system_name,
            permission_level='basic'
        )
        db.session.add(perm)
    
    perm.can_view = 1 if 'view' in permissions else 0
    perm.can_add = 1 if 'add' in permissions else 0
    perm.can_edit = 1 if 'edit' in permissions else 0
    perm.can_delete = 1 if 'delete' in permissions else 0
    perm.can_export = 1 if 'export' in permissions else 0
    perm.can_import = 1 if 'import' in permissions else 0
    perm.can_approve = 1 if 'approve' in permissions else 0
    perm.can_config = 1 if 'config' in permissions else 0
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '权限更新成功'})

@bp.route('/security/permissions/update', methods=['POST'])
@login_required
@admin_required
def permission_update():
    """更新权限"""
    role_id = request.form.get('role_id')
    module_id = request.form.get('module_id')
    action = request.form.get('action')  # view, add, edit, delete, export, import, audit, approve
    value = int(request.form.get('value', 0))
    
    perm = SysPermission.query.filter_by(role_id=role_id, module_id=module_id).first()
    
    if not perm:
        perm = SysPermission(role_id=role_id, module_id=module_id)
        db.session.add(perm)
    
    if action == 'view':
        perm.can_view = value
    elif action == 'add':
        perm.can_add = value
    elif action == 'edit':
        perm.can_edit = value
    elif action == 'delete':
        perm.can_delete = value
    elif action == 'export':
        perm.can_export = value
    elif action == 'import':
        perm.can_import = value
    elif action == 'audit':
        perm.can_audit = value
    elif action == 'approve':
        perm.can_approve = value
    elif action == 'data_scope':
        perm.data_scope = request.form.get('data_scope', 'self')
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '权限更新成功'})

@bp.route('/security/permissions/batch', methods=['POST'])
@login_required
@admin_required
def permission_batch():
    """批量设置权限"""
    role_id = request.form.get('role_id')
    module_id = request.form.get('module_id')
    
    perm = SysPermission.query.filter_by(role_id=role_id, module_id=module_id).first()
    
    if not perm:
        perm = SysPermission(role_id=role_id, module_id=module_id)
        db.session.add(perm)
    
    perm.can_view = 1
    perm.can_add = 1
    perm.can_edit = 1
    perm.can_delete = 1
    perm.can_export = 1
    perm.can_import = 1
    perm.can_audit = 1
    perm.can_approve = 1
    perm.data_scope = 'all'
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '批量权限设置成功'})

# ==================== 人员权限矩阵（按公司-系统） ====================

@bp.route('/security/person_permission_matrix')
@login_required
@admin_required
def person_permission_matrix():
    """人员权限矩阵（按人员-系统维度）"""
    employees = db.session.query(
        PersonSystemPermissionMatrix.emp_id,
        PersonSystemPermissionMatrix.emp_name,
        PersonSystemPermissionMatrix.dept_id,
        PersonSystemPermissionMatrix.dept_name
    ).distinct().all()
    
    systems = db.session.query(PersonSystemPermissionMatrix.system_name).distinct().all()
    systems = [s[0] for s in systems]
    
    permissions = PersonSystemPermissionMatrix.query.all()
    
    permission_dict = {}
    for perm in permissions:
        key = (perm.emp_id, perm.system_name)
        permission_dict[key] = {
            'can_view': perm.can_view,
            'can_add': perm.can_add,
            'can_edit': perm.can_edit,
            'can_delete': perm.can_delete,
            'can_export': perm.can_export,
            'can_import': perm.can_import,
            'can_approve': perm.can_approve,
            'can_config': perm.can_config
        }
    
    return render_template('security/person_permission_matrix.html',
                           employees=employees,
                           systems=systems,
                           permission_dict=permission_dict,
                           permission_count=len(permissions))

# ==================== 用户角色分配 ====================

@bp.route('/security/user_roles')
@login_required
@admin_required
def user_role_list():
    """用户角色列表"""
    from app.models import User
    users = User.query.all()
    roles = SysRole.query.filter_by(status=1).all()
    
    user_roles = {}
    for user in users:
        user_roles[user.id] = SysUserRole.query.filter_by(user_id=user.id).all()
    
    return render_template('security/user_role_list.html', 
                           users=users, 
                           roles=roles, 
                           user_roles=user_roles)

@bp.route('/security/user_roles/assign', methods=['POST'])
@login_required
@admin_required
def user_role_assign():
    """分配用户角色"""
    user_id = request.form.get('user_id')
    role_id = request.form.get('role_id')
    
    # 检查是否已分配
    existing = SysUserRole.query.filter_by(user_id=user_id, role_id=role_id).first()
    if existing:
        return jsonify({'success': False, 'message': '该角色已分配给此用户'})
    
    user_role = SysUserRole(user_id=user_id, role_id=role_id)
    db.session.add(user_role)
    db.session.commit()
    
    return jsonify({'success': True, 'message': '角色分配成功'})

@bp.route('/security/user_roles/remove', methods=['POST'])
@login_required
@admin_required
def user_role_remove():
    """移除用户角色"""
    user_id = request.form.get('user_id')
    role_id = request.form.get('role_id')
    
    user_role = SysUserRole.query.filter_by(user_id=user_id, role_id=role_id).first()
    if user_role:
        db.session.delete(user_role)
        db.session.commit()
    
    return jsonify({'success': True, 'message': '角色移除成功'})

# ==================== 涉密人员管理 ====================

@bp.route('/security/classified_personnel')
@login_required
@admin_required
def classified_personnel():
    items = ClassifiedPersonnel.query.all()
    return render_template('security/classified_personnel.html', items=items)

# ==================== 涉密介质管理 ====================

@bp.route('/security/classified_media')
@login_required
@admin_required
def classified_media():
    items = ClassifiedMedia.query.all()
    return render_template('security/classified_media.html', items=items)

# ==================== 安全区域管理 ====================

@bp.route('/security/security_zone')
@login_required
@admin_required
def security_zone():
    items = SecurityZone.query.all()
    return render_template('security/security_zone.html', items=items)

# ==================== 电子文件管理 ====================

@bp.route('/security/electronic_document')
@login_required
@admin_required
def electronic_document():
    items = ElectronicDocument.query.all()
    return render_template('security/electronic_document.html', items=items)

# ==================== 纸质文件管理 ====================

@bp.route('/security/paper_document')
@login_required
@admin_required
def paper_document():
    items = PaperDocument.query.all()
    return render_template('security/paper_document.html', items=items)

# ==================== 涉密人员 CRUD ====================

@bp.route('/security/classified_personnel/add', methods=['POST'])
@login_required
@admin_required
def add_classified_personnel():
    """添加涉密人员"""
    try:
        item = ClassifiedPersonnel(
            emp_id=request.form.get('emp_id'),
            emp_name=request.form.get('emp_name'),
            dept_id=safe_int(request.form.get('dept_id')),
            dept_name=request.form.get('dept_name'),
            position=request.form.get('position'),
            classification_level=request.form.get('classification_level'),
            training_record=request.form.get('training_record'),
            agreement_type=request.form.get('agreement_type'),
            agreement_sign_date=request.form.get('agreement_sign_date') or None,
            signing_date=request.form.get('signing_date') or None,
            expiration_date=request.form.get('expiration_date') or None,
            status=request.form.get('status', '有效'),
            remark=request.form.get('remark')
        )
        db.session.add(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '添加成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/security/classified_personnel/get/<int:id>')
@login_required
@admin_required
def get_classified_personnel(id):
    item = ClassifiedPersonnel.query.get_or_404(id)
    return jsonify({
        'id': item.id, 'emp_id': item.emp_id, 'emp_name': item.emp_name,
        'dept_id': item.dept_id, 'dept_name': item.dept_name,
        'position': item.position, 'classification_level': item.classification_level,
        'training_record': item.training_record, 'agreement_type': item.agreement_type,
        'agreement_sign_date': str(item.agreement_sign_date) if item.agreement_sign_date else '',
        'signing_date': str(item.signing_date) if item.signing_date else '',
        'expiration_date': str(item.expiration_date) if item.expiration_date else '',
        'status': item.status, 'remark': item.remark
    })

@bp.route('/security/classified_personnel/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_classified_personnel(id):
    item = ClassifiedPersonnel.query.get_or_404(id)
    try:
        item.emp_id = request.form.get('emp_id')
        item.emp_name = request.form.get('emp_name')
        item.dept_id = safe_int(request.form.get('dept_id'))
        item.dept_name = request.form.get('dept_name')
        item.position = request.form.get('position')
        item.classification_level = request.form.get('classification_level')
        item.training_record = request.form.get('training_record')
        item.agreement_type = request.form.get('agreement_type')
        item.agreement_sign_date = request.form.get('agreement_sign_date') or None
        item.signing_date = request.form.get('signing_date') or None
        item.expiration_date = request.form.get('expiration_date') or None
        item.status = request.form.get('status')
        item.remark = request.form.get('remark')
        db.session.commit()
        return jsonify({'success': True, 'message': '更新成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/security/classified_personnel/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_classified_personnel(id):
    item = ClassifiedPersonnel.query.get_or_404(id)
    try:
        db.session.delete(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# ==================== 涉密介质 CRUD ====================

@bp.route('/security/classified_media/add', methods=['POST'])
@login_required
@admin_required
def add_classified_media():
    """添加涉密介质"""
    try:
        item = ClassifiedMedia(
            media_id=request.form.get('media_id'),
            media_type=request.form.get('media_type'),
            brand_model=request.form.get('brand_model'),
            serial_no=request.form.get('serial_no'),
            classification=request.form.get('classification'),
            custodian_id=request.form.get('custodian_id'),
            custodian_name=request.form.get('custodian_name'),
            dept_id=safe_int(request.form.get('dept_id')),
            dept_name=request.form.get('dept_name'),
            purpose=request.form.get('purpose'),
            status=request.form.get('status', '在用'),
            remark=request.form.get('remark')
        )
        db.session.add(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '添加成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/security/classified_media/get/<int:id>')
@login_required
@admin_required
def get_classified_media(id):
    item = ClassifiedMedia.query.get_or_404(id)
    return jsonify({
        'id': item.id, 'media_id': item.media_id, 'media_type': item.media_type,
        'brand_model': item.brand_model, 'serial_no': item.serial_no,
        'classification': item.classification, 'custodian_id': item.custodian_id,
        'custodian_name': item.custodian_name, 'dept_id': item.dept_id,
        'dept_name': item.dept_name, 'purpose': item.purpose,
        'status': item.status, 'remark': item.remark
    })

@bp.route('/security/classified_media/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_classified_media(id):
    item = ClassifiedMedia.query.get_or_404(id)
    try:
        item.media_id = request.form.get('media_id')
        item.media_type = request.form.get('media_type')
        item.brand_model = request.form.get('brand_model')
        item.serial_no = request.form.get('serial_no')
        item.classification = request.form.get('classification')
        item.custodian_id = request.form.get('custodian_id')
        item.custodian_name = request.form.get('custodian_name')
        item.dept_id = safe_int(request.form.get('dept_id'))
        item.dept_name = request.form.get('dept_name')
        item.purpose = request.form.get('purpose')
        item.status = request.form.get('status')
        item.remark = request.form.get('remark')
        db.session.commit()
        return jsonify({'success': True, 'message': '更新成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/security/classified_media/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_classified_media(id):
    item = ClassifiedMedia.query.get_or_404(id)
    try:
        db.session.delete(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# ==================== 安全区域 CRUD ====================

@bp.route('/security/security_zone/add', methods=['POST'])
@login_required
@admin_required
def add_security_zone():
    """添加安全区域"""
    try:
        item = SecurityZone(
            zone_id=request.form.get('zone_id'),
            zone_name=request.form.get('zone_name'),
            zone_type=request.form.get('zone_type'),
            location=request.form.get('location'),
            manager_id=request.form.get('manager_id'),
            manager_name=request.form.get('manager_name'),
            dept_id=safe_int(request.form.get('dept_id')),
            dept_name=request.form.get('dept_name'),
            zone_level=request.form.get('zone_level'),
            status=request.form.get('status', '正常'),
            remark=request.form.get('remark')
        )
        db.session.add(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '添加成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/security/security_zone/get/<int:id>')
@login_required
@admin_required
def get_security_zone(id):
    item = SecurityZone.query.get_or_404(id)
    return jsonify({
        'id': item.id, 'zone_id': item.zone_id, 'zone_name': item.zone_name,
        'zone_type': item.zone_type, 'location': item.location,
        'manager_id': item.manager_id, 'manager_name': item.manager_name,
        'dept_id': item.dept_id, 'dept_name': item.dept_name,
        'zone_level': item.zone_level, 'status': item.status, 'remark': item.remark
    })

@bp.route('/security/security_zone/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_security_zone(id):
    item = SecurityZone.query.get_or_404(id)
    try:
        item.zone_id = request.form.get('zone_id')
        item.zone_name = request.form.get('zone_name')
        item.zone_type = request.form.get('zone_type')
        item.location = request.form.get('location')
        item.manager_id = request.form.get('manager_id')
        item.manager_name = request.form.get('manager_name')
        item.dept_id = safe_int(request.form.get('dept_id'))
        item.dept_name = request.form.get('dept_name')
        item.zone_level = request.form.get('zone_level')
        item.status = request.form.get('status')
        item.remark = request.form.get('remark')
        db.session.commit()
        return jsonify({'success': True, 'message': '更新成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/security/security_zone/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_security_zone(id):
    item = SecurityZone.query.get_or_404(id)
    try:
        db.session.delete(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# ==================== 电子文件 CRUD ====================

@bp.route('/security/electronic_document/add', methods=['POST'])
@login_required
@admin_required
def add_electronic_document():
    """添加电子文件"""
    try:
        item = ElectronicDocument(
            doc_id=request.form.get('doc_id'),
            doc_title=request.form.get('doc_title'),
            classification=request.form.get('classification'),
            file_format=request.form.get('file_format'),
            drafter_id=request.form.get('drafter_id'),
            drafter_name=request.form.get('drafter_name'),
            draft_dept=request.form.get('draft_dept'),
            storage_path=request.form.get('storage_path'),
            custodian_id=request.form.get('custodian_id'),
            custodian_name=request.form.get('custodian_name'),
            dept_id=safe_int(request.form.get('dept_id')),
            dept_name=request.form.get('dept_name'),
            doc_status=request.form.get('doc_status', '正常'),
            remark=request.form.get('remark')
        )
        db.session.add(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '添加成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/security/electronic_document/get/<int:id>')
@login_required
@admin_required
def get_electronic_document(id):
    item = ElectronicDocument.query.get_or_404(id)
    return jsonify({
        'id': item.id, 'doc_id': item.doc_id, 'doc_title': item.doc_title,
        'classification': item.classification, 'file_format': item.file_format,
        'drafter_id': item.drafter_id, 'drafter_name': item.drafter_name,
        'draft_dept': item.draft_dept, 'storage_path': item.storage_path,
        'custodian_id': item.custodian_id, 'custodian_name': item.custodian_name,
        'dept_id': item.dept_id, 'dept_name': item.dept_name,
        'doc_status': item.doc_status, 'remark': item.remark
    })

@bp.route('/security/electronic_document/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_electronic_document(id):
    item = ElectronicDocument.query.get_or_404(id)
    try:
        item.doc_id = request.form.get('doc_id')
        item.doc_title = request.form.get('doc_title')
        item.classification = request.form.get('classification')
        item.file_format = request.form.get('file_format')
        item.drafter_id = request.form.get('drafter_id')
        item.drafter_name = request.form.get('drafter_name')
        item.draft_dept = request.form.get('draft_dept')
        item.storage_path = request.form.get('storage_path')
        item.custodian_id = request.form.get('custodian_id')
        item.custodian_name = request.form.get('custodian_name')
        item.dept_id = safe_int(request.form.get('dept_id'))
        item.dept_name = request.form.get('dept_name')
        item.doc_status = request.form.get('doc_status')
        item.remark = request.form.get('remark')
        db.session.commit()
        return jsonify({'success': True, 'message': '更新成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/security/electronic_document/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_electronic_document(id):
    item = ElectronicDocument.query.get_or_404(id)
    try:
        db.session.delete(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# ==================== 纸质文件 CRUD ====================

@bp.route('/security/paper_document/add', methods=['POST'])
@login_required
@admin_required
def add_paper_document():
    """添加纸质文件"""
    try:
        item = PaperDocument(
            doc_id=request.form.get('doc_id'),
            doc_title=request.form.get('doc_title'),
            classification=request.form.get('classification'),
            copies=safe_int(request.form.get('copies')),
            pages=safe_int(request.form.get('pages')),
            drafter_id=request.form.get('drafter_id'),
            drafter_name=request.form.get('drafter_name'),
            holder_id=request.form.get('holder_id'),
            holder_name=request.form.get('holder_name'),
            storage_location=request.form.get('storage_location'),
            custodian_id=request.form.get('custodian_id'),
            custodian_name=request.form.get('custodian_name'),
            dept_id=safe_int(request.form.get('dept_id')),
            dept_name=request.form.get('dept_name'),
            doc_status=request.form.get('doc_status', '正常'),
            remark=request.form.get('remark')
        )
        db.session.add(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '添加成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/security/paper_document/get/<int:id>')
@login_required
@admin_required
def get_paper_document(id):
    item = PaperDocument.query.get_or_404(id)
    return jsonify({
        'id': item.id, 'doc_id': item.doc_id, 'doc_title': item.doc_title,
        'classification': item.classification, 'copies': item.copies,
        'pages': item.pages, 'drafter_id': item.drafter_id,
        'drafter_name': item.drafter_name, 'holder_id': item.holder_id,
        'holder_name': item.holder_name, 'storage_location': item.storage_location,
        'custodian_id': item.custodian_id, 'custodian_name': item.custodian_name,
        'dept_id': item.dept_id, 'dept_name': item.dept_name,
        'doc_status': item.doc_status, 'remark': item.remark
    })

@bp.route('/security/paper_document/edit/<int:id>', methods=['POST'])
@login_required
@admin_required
def edit_paper_document(id):
    item = PaperDocument.query.get_or_404(id)
    try:
        item.doc_id = request.form.get('doc_id')
        item.doc_title = request.form.get('doc_title')
        item.classification = request.form.get('classification')
        item.copies = safe_int(request.form.get('copies'))
        item.pages = safe_int(request.form.get('pages'))
        item.drafter_id = request.form.get('drafter_id')
        item.drafter_name = request.form.get('drafter_name')
        item.holder_id = request.form.get('holder_id')
        item.holder_name = request.form.get('holder_name')
        item.storage_location = request.form.get('storage_location')
        item.custodian_id = request.form.get('custodian_id')
        item.custodian_name = request.form.get('custodian_name')
        item.dept_id = safe_int(request.form.get('dept_id'))
        item.dept_name = request.form.get('dept_name')
        item.doc_status = request.form.get('doc_status')
        item.remark = request.form.get('remark')
        db.session.commit()
        return jsonify({'success': True, 'message': '更新成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

@bp.route('/security/paper_document/delete/<int:id>', methods=['POST'])
@login_required
@admin_required
def delete_paper_document(id):
    item = PaperDocument.query.get_or_404(id)
    try:
        db.session.delete(item)
        db.session.commit()
        return jsonify({'success': True, 'message': '删除成功'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)})

# ==================== 涉密人员导入导出 ====================

@bp.route('/security/classified_personnel/export')
@login_required
@admin_required
def export_classified_personnel():
    """导出涉密人员数据"""
    items = ClassifiedPersonnel.query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '涉密人员'
    
    headers = ['工号', '姓名', '部门ID', '部门名称', '岗位', '涉密等级', '培训记录', 
               '协议类型', '协议签订日期', '状态', '签订日期', '到期日期', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.emp_id or '').border = thin_border
        ws.cell(row=row, column=2, value=item.emp_name or '').border = thin_border
        ws.cell(row=row, column=3, value=item.dept_id or '').border = thin_border
        ws.cell(row=row, column=4, value=item.dept_name or '').border = thin_border
        ws.cell(row=row, column=5, value=item.position or '').border = thin_border
        ws.cell(row=row, column=6, value=item.classification_level or '').border = thin_border
        ws.cell(row=row, column=7, value=item.training_record or '').border = thin_border
        ws.cell(row=row, column=8, value=item.agreement_type or '').border = thin_border
        ws.cell(row=row, column=9, value=str(item.agreement_sign_date) if item.agreement_sign_date else '').border = thin_border
        ws.cell(row=row, column=10, value=item.status or '').border = thin_border
        ws.cell(row=row, column=11, value=str(item.signing_date) if item.signing_date else '').border = thin_border
        ws.cell(row=row, column=12, value=str(item.expiration_date) if item.expiration_date else '').border = thin_border
        ws.cell(row=row, column=13, value=item.remark or '').border = thin_border
    
    column_widths = [12, 12, 12, 20, 15, 12, 15, 12, 15, 10, 12, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'涉密人员_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@bp.route('/security/classified_personnel/import', methods=['POST'])
@login_required
@admin_required
def import_classified_personnel():
    """导入涉密人员数据"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '不支持的文件格式'})
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        success_count = 0
        error_count = 0
        
        for row_num in range(2, ws.max_row + 1):
            try:
                emp_id = str(ws.cell(row=row_num, column=1).value or '').strip()
                emp_name = str(ws.cell(row=row_num, column=2).value or '').strip()
                
                if not emp_id or not emp_name:
                    error_count += 1
                    continue
                
                existing = ClassifiedPersonnel.query.filter_by(emp_id=emp_id).first()
                
                if existing:
                    existing.emp_name = emp_name
                    existing.dept_id = safe_int(ws.cell(row=row_num, column=3).value)
                    existing.dept_name = str(ws.cell(row=row_num, column=4).value or '').strip()
                    existing.position = str(ws.cell(row=row_num, column=5).value or '').strip()
                    existing.classification_level = str(ws.cell(row=row_num, column=6).value or '').strip()
                    existing.training_record = str(ws.cell(row=row_num, column=7).value or '').strip()
                    existing.agreement_type = str(ws.cell(row=row_num, column=8).value or '').strip()
                    existing.agreement_sign_date = ws.cell(row=row_num, column=9).value
                    existing.status = str(ws.cell(row=row_num, column=10).value or '').strip()
                    existing.signing_date = ws.cell(row=row_num, column=11).value
                    existing.expiration_date = ws.cell(row=row_num, column=12).value
                    existing.remark = str(ws.cell(row=row_num, column=13).value or '').strip()
                else:
                    item = ClassifiedPersonnel(
                        emp_id=emp_id,
                        emp_name=emp_name,
                        dept_id=safe_int(ws.cell(row=row_num, column=3).value),
                        dept_name=str(ws.cell(row=row_num, column=4).value or '').strip(),
                        position=str(ws.cell(row=row_num, column=5).value or '').strip(),
                        classification_level=str(ws.cell(row=row_num, column=6).value or '').strip(),
                        training_record=str(ws.cell(row=row_num, column=7).value or '').strip(),
                        agreement_type=str(ws.cell(row=row_num, column=8).value or '').strip(),
                        agreement_sign_date=ws.cell(row=row_num, column=9).value,
                        status=str(ws.cell(row=row_num, column=10).value or '').strip(),
                        signing_date=ws.cell(row=row_num, column=11).value,
                        expiration_date=ws.cell(row=row_num, column=12).value,
                        remark=str(ws.cell(row=row_num, column=13).value or '').strip()
                    )
                    db.session.add(item)
                
                success_count += 1
            except Exception:
                error_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'导入完成：成功{success_count}条，失败{error_count}条'
        })
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败：{str(e)}'})

@bp.route('/security/classified_personnel/template')
@login_required
@admin_required
def download_classified_personnel_template():
    """下载涉密人员导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '涉密人员导入模板'
    
    headers = ['工号', '姓名', '部门ID', '部门名称', '岗位', '涉密等级', '培训记录', 
               '协议类型', '协议签订日期', '状态', '签订日期', '到期日期', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    sample_data = ['E001', '张三', 1, '技术部', '工程师', '机密', '已培训', 
                   '保密协议', '2024-01-01', '在职', '2024-01-01', '2025-01-01', '备注']
    for col, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col, value=value).border = thin_border
    
    column_widths = [12, 12, 12, 20, 15, 12, 15, 12, 15, 10, 12, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='涉密人员导入模板.xlsx'
    )

# ==================== 涉密介质导入导出 ====================

@bp.route('/security/classified_media/export')
@login_required
@admin_required
def export_classified_media():
    """导出涉密介质数据"""
    items = ClassifiedMedia.query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '涉密介质'
    
    headers = ['介质编号', '介质类型', '品牌型号', '序列号', '涉密等级', '保管人工号', 
               '保管人姓名', '部门ID', '部门名称', '用途', '状态', '介质数量', '容量',
               '负责人姓名', '负责人工号', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.media_id or '').border = thin_border
        ws.cell(row=row, column=2, value=item.media_type or '').border = thin_border
        ws.cell(row=row, column=3, value=item.brand_model or '').border = thin_border
        ws.cell(row=row, column=4, value=item.serial_no or '').border = thin_border
        ws.cell(row=row, column=5, value=item.classification or '').border = thin_border
        ws.cell(row=row, column=6, value=item.custodian_id or '').border = thin_border
        ws.cell(row=row, column=7, value=item.custodian_name or '').border = thin_border
        ws.cell(row=row, column=8, value=item.dept_id or '').border = thin_border
        ws.cell(row=row, column=9, value=item.dept_name or '').border = thin_border
        ws.cell(row=row, column=10, value=item.purpose or '').border = thin_border
        ws.cell(row=row, column=11, value=item.status or '').border = thin_border
        ws.cell(row=row, column=12, value=item.media_number or '').border = thin_border
        ws.cell(row=row, column=13, value=item.capacity or '').border = thin_border
        ws.cell(row=row, column=14, value=item.responsible_name or '').border = thin_border
        ws.cell(row=row, column=15, value=item.responsible_emp_id or '').border = thin_border
        ws.cell(row=row, column=16, value=item.remark or '').border = thin_border
    
    column_widths = [15, 12, 20, 20, 12, 12, 12, 12, 20, 15, 10, 10, 10, 12, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'涉密介质_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@bp.route('/security/classified_media/import', methods=['POST'])
@login_required
@admin_required
def import_classified_media():
    """导入涉密介质数据"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '请选择有效的Excel文件'})
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        success_count = 0
        error_count = 0
        
        for row_num in range(2, ws.max_row + 1):
            try:
                media_id = str(ws.cell(row=row_num, column=1).value or '').strip()
                if not media_id:
                    error_count += 1
                    continue
                
                existing = ClassifiedMedia.query.filter_by(media_id=media_id).first()
                
                data = {
                    'media_id': media_id,
                    'media_type': str(ws.cell(row=row_num, column=2).value or '').strip(),
                    'brand_model': str(ws.cell(row=row_num, column=3).value or '').strip(),
                    'serial_no': str(ws.cell(row=row_num, column=4).value or '').strip(),
                    'classification': str(ws.cell(row=row_num, column=5).value or '').strip(),
                    'custodian_id': str(ws.cell(row=row_num, column=6).value or '').strip(),
                    'custodian_name': str(ws.cell(row=row_num, column=7).value or '').strip(),
                    'dept_id': safe_int(ws.cell(row=row_num, column=8).value),
                    'dept_name': str(ws.cell(row=row_num, column=9).value or '').strip(),
                    'purpose': str(ws.cell(row=row_num, column=10).value or '').strip(),
                    'status': str(ws.cell(row=row_num, column=11).value or '').strip(),
                    'media_number': str(ws.cell(row=row_num, column=12).value or '').strip(),
                    'capacity': str(ws.cell(row=row_num, column=13).value or '').strip(),
                    'responsible_name': str(ws.cell(row=row_num, column=14).value or '').strip(),
                    'responsible_emp_id': str(ws.cell(row=row_num, column=15).value or '').strip(),
                    'remark': str(ws.cell(row=row_num, column=16).value or '').strip()
                }
                
                if existing:
                    for key, value in data.items():
                        setattr(existing, key, value)
                else:
                    item = ClassifiedMedia(**data)
                    db.session.add(item)
                
                success_count += 1
            except Exception:
                error_count += 1
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'导入完成：成功{success_count}条，失败{error_count}条'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败：{str(e)}'})

@bp.route('/security/classified_media/template')
@login_required
@admin_required
def download_classified_media_template():
    """下载涉密介质导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '涉密介质导入模板'
    
    headers = ['介质编号', '介质类型', '品牌型号', '序列号', '涉密等级', '保管人工号', 
               '保管人姓名', '部门ID', '部门名称', '用途', '状态', '介质数量', '容量',
               '负责人姓名', '负责人工号', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    sample = ['M001', 'U盘', '金士顿DT100', 'SN123456', '机密', 'E001', '张三', 'D001', '技术部', '数据存储', '在用', '1', '32GB', '李四', 'E002', '备注']
    for col, value in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=value).border = thin_border
    
    column_widths = [15, 12, 20, 20, 12, 12, 12, 12, 20, 15, 10, 10, 10, 12, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name='涉密介质导入模板.xlsx')

# ==================== 安全区域导入导出 ====================

@bp.route('/security/security_zone/export')
@login_required
@admin_required
def export_security_zone():
    """导出安全区域数据"""
    items = SecurityZone.query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '安全区域'
    
    headers = ['区域编号', '区域名称', '区域类型', '位置', '管理人工号', '管理人姓名', 
               '部门ID', '部门名称', '状态', '区域代码', '区域等级', '负责人姓名', '负责人工号', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.zone_id or '').border = thin_border
        ws.cell(row=row, column=2, value=item.zone_name or '').border = thin_border
        ws.cell(row=row, column=3, value=item.zone_type or '').border = thin_border
        ws.cell(row=row, column=4, value=item.location or '').border = thin_border
        ws.cell(row=row, column=5, value=item.manager_id or '').border = thin_border
        ws.cell(row=row, column=6, value=item.manager_name or '').border = thin_border
        ws.cell(row=row, column=7, value=item.dept_id or '').border = thin_border
        ws.cell(row=row, column=8, value=item.dept_name or '').border = thin_border
        ws.cell(row=row, column=9, value=item.status or '').border = thin_border
        ws.cell(row=row, column=10, value=item.zone_code or '').border = thin_border
        ws.cell(row=row, column=11, value=item.zone_level or '').border = thin_border
        ws.cell(row=row, column=12, value=item.responsible_name or '').border = thin_border
        ws.cell(row=row, column=13, value=item.responsible_emp_id or '').border = thin_border
        ws.cell(row=row, column=14, value=item.remark or '').border = thin_border
    
    column_widths = [12, 20, 12, 20, 12, 12, 12, 20, 10, 12, 10, 12, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'安全区域_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

@bp.route('/security/security_zone/import', methods=['POST'])
@login_required
@admin_required
def import_security_zone():
    """导入安全区域数据"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '请选择有效的Excel文件'})
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        success_count = 0
        error_count = 0
        
        for row_num in range(2, ws.max_row + 1):
            try:
                zone_id = str(ws.cell(row=row_num, column=1).value or '').strip()
                if not zone_id:
                    error_count += 1
                    continue
                
                existing = SecurityZone.query.filter_by(zone_id=zone_id).first()
                
                data = {
                    'zone_id': zone_id,
                    'zone_name': str(ws.cell(row=row_num, column=2).value or '').strip(),
                    'zone_type': str(ws.cell(row=row_num, column=3).value or '').strip(),
                    'location': str(ws.cell(row=row_num, column=4).value or '').strip(),
                    'manager_id': str(ws.cell(row=row_num, column=5).value or '').strip(),
                    'manager_name': str(ws.cell(row=row_num, column=6).value or '').strip(),
                    'dept_id': safe_int(ws.cell(row=row_num, column=7).value),
                    'dept_name': str(ws.cell(row=row_num, column=8).value or '').strip(),
                    'status': str(ws.cell(row=row_num, column=9).value or '').strip(),
                    'zone_code': str(ws.cell(row=row_num, column=10).value or '').strip(),
                    'zone_level': str(ws.cell(row=row_num, column=11).value or '').strip(),
                    'responsible_name': str(ws.cell(row=row_num, column=12).value or '').strip(),
                    'responsible_emp_id': str(ws.cell(row=row_num, column=13).value or '').strip(),
                    'remark': str(ws.cell(row=row_num, column=14).value or '').strip()
                }
                
                if existing:
                    for key, value in data.items():
                        setattr(existing, key, value)
                else:
                    item = SecurityZone(**data)
                    db.session.add(item)
                
                success_count += 1
            except Exception:
                error_count += 1
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'导入完成：成功{success_count}条，失败{error_count}条'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败：{str(e)}'})

@bp.route('/security/security_zone/template')
@login_required
@admin_required
def download_security_zone_template():
    """下载安全区域导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '安全区域导入模板'
    
    headers = ['区域编号', '区域名称', '区域类型', '位置', '管理人工号', '管理人姓名', 
               '部门ID', '部门名称', '状态', '区域代码', '区域等级', '负责人姓名', '负责人工号', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    sample = ['Z001', '机房A区', '机房', '一楼东侧', 'E001', '张三', 'D001', '技术部', '正常', 'A001', '一级', '李四', 'E002', '备注']
    for col, value in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=value).border = thin_border
    
    column_widths = [12, 20, 12, 20, 12, 12, 12, 20, 10, 12, 10, 12, 12, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name='安全区域导入模板.xlsx')

# ==================== 电子文件导入导出 ====================

@bp.route('/security/electronic_document/export')
@login_required
@admin_required
def export_electronic_document():
    """导出电子文件数据"""
    items = ElectronicDocument.query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '电子文件'
    
    headers = ['文件编号', '文件标题', '涉密等级', '文件格式', '起草人工号', '起草人姓名', 
               '起草部门', '存储路径', '保管人工号', '保管人姓名', '部门ID', '部门名称', 
               '文件状态', '文件编号', '文件等级', '负责人姓名', '负责人工号', '文件路径', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.doc_id or '').border = thin_border
        ws.cell(row=row, column=2, value=item.doc_title or '').border = thin_border
        ws.cell(row=row, column=3, value=item.classification or '').border = thin_border
        ws.cell(row=row, column=4, value=item.file_format or '').border = thin_border
        ws.cell(row=row, column=5, value=item.drafter_id or '').border = thin_border
        ws.cell(row=row, column=6, value=item.drafter_name or '').border = thin_border
        ws.cell(row=row, column=7, value=item.draft_dept or '').border = thin_border
        ws.cell(row=row, column=8, value=item.storage_path or '').border = thin_border
        ws.cell(row=row, column=9, value=item.custodian_id or '').border = thin_border
        ws.cell(row=row, column=10, value=item.custodian_name or '').border = thin_border
        ws.cell(row=row, column=11, value=item.dept_id or '').border = thin_border
        ws.cell(row=row, column=12, value=item.dept_name or '').border = thin_border
        ws.cell(row=row, column=13, value=item.doc_status or '').border = thin_border
        ws.cell(row=row, column=14, value=item.doc_number or '').border = thin_border
        ws.cell(row=row, column=15, value=item.doc_level or '').border = thin_border
        ws.cell(row=row, column=16, value=item.responsible_name or '').border = thin_border
        ws.cell(row=row, column=17, value=item.responsible_emp_id or '').border = thin_border
        ws.cell(row=row, column=18, value=item.file_path or '').border = thin_border
        ws.cell(row=row, column=19, value=item.remark or '').border = thin_border
    
    column_widths = [12, 30, 12, 10, 12, 12, 15, 30, 12, 12, 12, 20, 10, 12, 10, 12, 12, 30, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'电子文件_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

@bp.route('/security/electronic_document/import', methods=['POST'])
@login_required
@admin_required
def import_electronic_document():
    """导入电子文件数据"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '请选择有效的Excel文件'})
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        success_count = 0
        error_count = 0
        
        for row_num in range(2, ws.max_row + 1):
            try:
                doc_id = str(ws.cell(row=row_num, column=1).value or '').strip()
                if not doc_id:
                    error_count += 1
                    continue
                
                existing = ElectronicDocument.query.filter_by(doc_id=doc_id).first()
                
                data = {
                    'doc_id': doc_id,
                    'doc_title': str(ws.cell(row=row_num, column=2).value or '').strip(),
                    'classification': str(ws.cell(row=row_num, column=3).value or '').strip(),
                    'file_format': str(ws.cell(row=row_num, column=4).value or '').strip(),
                    'drafter_id': str(ws.cell(row=row_num, column=5).value or '').strip(),
                    'drafter_name': str(ws.cell(row=row_num, column=6).value or '').strip(),
                    'draft_dept': str(ws.cell(row=row_num, column=7).value or '').strip(),
                    'storage_path': str(ws.cell(row=row_num, column=8).value or '').strip(),
                    'custodian_id': str(ws.cell(row=row_num, column=9).value or '').strip(),
                    'custodian_name': str(ws.cell(row=row_num, column=10).value or '').strip(),
                    'dept_id': safe_int(ws.cell(row=row_num, column=11).value),
                    'dept_name': str(ws.cell(row=row_num, column=12).value or '').strip(),
                    'doc_status': str(ws.cell(row=row_num, column=13).value or '').strip(),
                    'doc_number': str(ws.cell(row=row_num, column=14).value or '').strip(),
                    'doc_level': str(ws.cell(row=row_num, column=15).value or '').strip(),
                    'responsible_name': str(ws.cell(row=row_num, column=16).value or '').strip(),
                    'responsible_emp_id': str(ws.cell(row=row_num, column=17).value or '').strip(),
                    'file_path': str(ws.cell(row=row_num, column=18).value or '').strip(),
                    'remark': str(ws.cell(row=row_num, column=19).value or '').strip()
                }
                
                if existing:
                    for key, value in data.items():
                        setattr(existing, key, value)
                else:
                    item = ElectronicDocument(**data)
                    db.session.add(item)
                
                success_count += 1
            except Exception:
                error_count += 1
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'导入完成：成功{success_count}条，失败{error_count}条'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败：{str(e)}'})

@bp.route('/security/electronic_document/template')
@login_required
@admin_required
def download_electronic_document_template():
    """下载电子文件导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '电子文件导入模板'
    
    headers = ['文件编号', '文件标题', '涉密等级', '文件格式', '起草人工号', '起草人姓名', 
               '起草部门', '存储路径', '保管人工号', '保管人姓名', '部门ID', '部门名称', 
               '文件状态', '文件编号', '文件等级', '负责人姓名', '负责人工号', '文件路径', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    sample = ['D001', '设计文档', '机密', 'docx', 'E001', '张三', '技术部', '/docs/2024', 'E002', '李四', 'D001', '技术部', '正常', 'DOC001', '一级', '王五', 'E003', '/files/doc001.docx', '备注']
    for col, value in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=value).border = thin_border
    
    column_widths = [12, 30, 12, 10, 12, 12, 15, 30, 12, 12, 12, 20, 10, 12, 10, 12, 12, 30, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name='电子文件导入模板.xlsx')

# ==================== 纸质文件导入导出 ====================

@bp.route('/security/paper_document/export')
@login_required
@admin_required
def export_paper_document():
    """导出纸质文件数据"""
    items = PaperDocument.query.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '纸质文件'
    
    headers = ['文件编号', '文件标题', '涉密等级', '份数', '页数', '起草人工号', '起草人姓名', 
               '持有工号', '持有人姓名', '存放位置', '保管人工号', '保管人姓名', '部门ID', '部门名称', 
               '文件状态', '文件编号', '文件等级', '负责人姓名', '负责人工号', '数量', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.doc_id or '').border = thin_border
        ws.cell(row=row, column=2, value=item.doc_title or '').border = thin_border
        ws.cell(row=row, column=3, value=item.classification or '').border = thin_border
        ws.cell(row=row, column=4, value=item.copies or '').border = thin_border
        ws.cell(row=row, column=5, value=item.pages or '').border = thin_border
        ws.cell(row=row, column=6, value=item.drafter_id or '').border = thin_border
        ws.cell(row=row, column=7, value=item.drafter_name or '').border = thin_border
        ws.cell(row=row, column=8, value=item.holder_id or '').border = thin_border
        ws.cell(row=row, column=9, value=item.holder_name or '').border = thin_border
        ws.cell(row=row, column=10, value=item.storage_location or '').border = thin_border
        ws.cell(row=row, column=11, value=item.custodian_id or '').border = thin_border
        ws.cell(row=row, column=12, value=item.custodian_name or '').border = thin_border
        ws.cell(row=row, column=13, value=item.dept_id or '').border = thin_border
        ws.cell(row=row, column=14, value=item.dept_name or '').border = thin_border
        ws.cell(row=row, column=15, value=item.doc_status or '').border = thin_border
        ws.cell(row=row, column=16, value=item.doc_number or '').border = thin_border
        ws.cell(row=row, column=17, value=item.doc_level or '').border = thin_border
        ws.cell(row=row, column=18, value=item.responsible_name or '').border = thin_border
        ws.cell(row=row, column=19, value=item.responsible_emp_id or '').border = thin_border
        ws.cell(row=row, column=20, value=item.quantity or '').border = thin_border
        ws.cell(row=row, column=21, value=item.remark or '').border = thin_border
    
    column_widths = [12, 30, 12, 8, 8, 12, 12, 12, 12, 20, 12, 12, 12, 20, 10, 12, 10, 12, 12, 8, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f'纸质文件_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')

@bp.route('/security/paper_document/import', methods=['POST'])
@login_required
@admin_required
def import_paper_document():
    """导入纸质文件数据"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    file = request.files['file']
    if file.filename == '' or not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '请选择有效的Excel文件'})
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        success_count = 0
        error_count = 0
        
        for row_num in range(2, ws.max_row + 1):
            try:
                doc_id = str(ws.cell(row=row_num, column=1).value or '').strip()
                if not doc_id:
                    error_count += 1
                    continue
                
                existing = PaperDocument.query.filter_by(doc_id=doc_id).first()
                
                data = {
                    'doc_id': doc_id,
                    'doc_title': str(ws.cell(row=row_num, column=2).value or '').strip(),
                    'classification': str(ws.cell(row=row_num, column=3).value or '').strip(),
                    'copies': safe_int(ws.cell(row=row_num, column=4).value),
                    'pages': safe_int(ws.cell(row=row_num, column=5).value),
                    'drafter_id': str(ws.cell(row=row_num, column=6).value or '').strip(),
                    'drafter_name': str(ws.cell(row=row_num, column=7).value or '').strip(),
                    'holder_id': str(ws.cell(row=row_num, column=8).value or '').strip(),
                    'holder_name': str(ws.cell(row=row_num, column=9).value or '').strip(),
                    'storage_location': str(ws.cell(row=row_num, column=10).value or '').strip(),
                    'custodian_id': str(ws.cell(row=row_num, column=11).value or '').strip(),
                    'custodian_name': str(ws.cell(row=row_num, column=12).value or '').strip(),
                    'dept_id': safe_int(ws.cell(row=row_num, column=13).value),
                    'dept_name': str(ws.cell(row=row_num, column=14).value or '').strip(),
                    'doc_status': str(ws.cell(row=row_num, column=15).value or '').strip(),
                    'doc_number': str(ws.cell(row=row_num, column=16).value or '').strip(),
                    'doc_level': str(ws.cell(row=row_num, column=17).value or '').strip(),
                    'responsible_name': str(ws.cell(row=row_num, column=18).value or '').strip(),
                    'responsible_emp_id': str(ws.cell(row=row_num, column=19).value or '').strip(),
                    'quantity': safe_int(ws.cell(row=row_num, column=20).value),
                    'remark': str(ws.cell(row=row_num, column=21).value or '').strip()
                }
                
                if existing:
                    for key, value in data.items():
                        setattr(existing, key, value)
                else:
                    item = PaperDocument(**data)
                    db.session.add(item)
                
                success_count += 1
            except Exception:
                error_count += 1
        
        db.session.commit()
        return jsonify({'success': True, 'message': f'导入完成：成功{success_count}条，失败{error_count}条'})
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败：{str(e)}'})

@bp.route('/security/paper_document/template')
@login_required
@admin_required
def download_paper_document_template():
    """下载纸质文件导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '纸质文件导入模板'
    
    headers = ['文件编号', '文件标题', '涉密等级', '份数', '页数', '起草人工号', '起草人姓名', 
               '持有工号', '持有人姓名', '存放位置', '保管人工号', '保管人姓名', '部门ID', '部门名称', 
               '文件状态', '文件编号', '文件等级', '负责人姓名', '负责人工号', '数量', '备注']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    sample = ['D001', '合同文件', '机密', '2', '10', 'E001', '张三', 'E002', '李四', '档案室A柜', 'E003', '王五', 'D001', '技术部', '正常', 'DOC001', '一级', '赵六', 'E004', '1', '备注']
    for col, value in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=value).border = thin_border
    
    column_widths = [12, 30, 12, 8, 8, 12, 12, 12, 12, 20, 12, 12, 12, 20, 10, 12, 10, 12, 12, 8, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name='纸质文件导入模板.xlsx')
