from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from app import db
from app.models import Personnel
from app.forms import PersonnelForm
from app.decorators import department_permission_required
from datetime import datetime
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

bp = Blueprint('personnel', __name__)

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@bp.route('/personnel')
@department_permission_required
def personnel_list():
    search = request.args.get('search', '').strip()[:100] if request.args.get('search') else None
    page = max(request.args.get('page', 1, type=int), 1)
    per_page = 50

    query = Personnel.query

    if search:
        query = query.filter(
            Personnel.emp_name.contains(search) |
            Personnel.emp_id.contains(search) |
            Personnel.dept_full_name.contains(search) |
            Personnel.position.contains(search) |
            Personnel.phone_number.contains(search)
        )

    if current_user.department_access:
        from app.decorators import get_user_accessible_departments, get_department_codes_recursive
        accessible_departments = get_user_accessible_departments()
        if accessible_departments:
            department_codes = []
            for dept_id in accessible_departments:
                department_codes.extend(get_department_codes_recursive(dept_id))
            query = query.filter(Personnel.dept_code.in_(department_codes))

    query = query.order_by(Personnel.id)
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    personnel = pagination.items
    return render_template('personnel/personnel.html', personnel=personnel, search=search, pagination=pagination)

@bp.route('/personnel/new', methods=['GET', 'POST'])
@department_permission_required
def new_personnel():
    form = PersonnelForm()
    if form.validate_on_submit():
        try:
            person = Personnel(
                # 部门信息
                company_dept=form.company_dept.data,
                factory_dept=form.factory_dept.data,
                dept_code=form.dept_code.data,
                dept_full_name=form.dept_full_name.data,
                business_unit=form.business_unit.data,
                first_level_dept_abbr=form.first_level_dept_abbr.data,
                org_category=form.org_category.data,
                job_family=form.job_family.data,
                dept_level1=form.dept_level1.data,
                dept_level2=form.dept_level2.data,
                dept_level3=form.dept_level3.data,
                dept_level4=form.dept_level4.data,
                leaf_dept=form.leaf_dept.data,
                region_org=form.region_org.data,
                work_location=form.work_location.data,
                
                # 基本信息
                emp_id=form.emp_id.data,
                emp_name=form.emp_name.data,
                emp_status=form.emp_status.data,
                emp_gender=form.emp_gender.data,
                emp_type=form.emp_type.data,
                group_hire_date=form.group_hire_date.data,
                hire_date=form.hire_date.data,
                is_probation=form.is_probation.data,
                probation_end_date=form.probation_end_date.data,
                regular_date=form.regular_date.data,
                
                # 职位信息
                position=form.position.data,
                position_category=form.position_category.data,
                position_grade=form.position_grade.data,
                job_type=form.job_type.data,
                job_category=form.job_category.data,
                job_sequence=form.job_sequence.data,
                job_title=form.job_title.data,
                acting_appointment=form.acting_appointment.data,
                project_job=form.project_job.data,
                job_rank=form.job_rank.data,
                daily_salary_grade=form.daily_salary_grade.data,
                qualification_grade=form.qualification_grade.data,
                professional_grade=form.professional_grade.data,
                job_level=form.job_level.data,
                project_job_level=form.project_job_level.data,
                talent_tag=form.talent_tag.data,
                is_mentor=form.is_mentor.data,
                
                # 汇报关系
                report_superior_id=form.report_superior_id.data,
                report_superior=form.report_superior.data,
                dept_head_id=form.dept_head_id.data,
                dept_head=form.dept_head.data,
                
                # 个人信息
                nationality=form.nationality.data,
                birthplace=form.birthplace.data,
                ethnicity=form.ethnicity.data,
                marital_status=form.marital_status.data,
                political_status=form.political_status.data,
                
                # 教育信息
                highest_education_type=form.highest_education_type.data,
                highest_education_mode=form.highest_education_mode.data,
                highest_education=form.highest_education.data,
                school=form.school.data,
                major=form.major.data,
                education_start_date=form.education_start_date.data,
                education_end_date=form.education_end_date.data,
                first_education=form.first_education.data,
                second_education=form.second_education.data,
                
                # 身份信息
                birth_date=form.birth_date.data,
                age=form.age.data,
                id_type=form.id_type.data,
                id_number=form.id_number.data,
                id_address=form.id_address.data,
                id_expiry=form.id_expiry.data,
                
                # 合同信息
                legal_entity=form.legal_entity.data,
                contract_entity=form.contract_entity.data,
                contract_type=form.contract_type.data,
                contract_start_date=form.contract_start_date.data,
                contract_end_date=form.contract_end_date.data,
                labor_company=form.labor_company.data,
                
                # 联系信息
                current_residence=form.current_residence.data,
                registered_residence=form.registered_residence.data,
                phone_number=form.phone_number.data,
                emergency_contact_name=form.emergency_contact_name.data,
                emergency_contact_phone=form.emergency_contact_phone.data,
                
                # 财务信息
                bank_name=form.bank_name.data,
                bank_branch=form.bank_branch.data,
                bank_account=form.bank_account.data,
                financial_code=form.financial_code.data,
                
                # 其他信息
                union_member=form.union_member.data,
                is_veteran=form.is_veteran.data,
                initial_social_insurance_unit=form.initial_social_insurance_unit.data,
                has_fuel_card=form.has_fuel_card.data,
                
                # 离职信息
                resignation_date=form.resignation_date.data,
                last_pay_date=form.last_pay_date.data,
                resignation_type=form.resignation_type.data,
                resignation_reason=form.resignation_reason.data,
                is_blacklisted=form.is_blacklisted.data,
                resignation_reason_detail=form.resignation_reason_detail.data,
                resignation_interview=form.resignation_interview.data
            )
            db.session.add(person)
            db.session.commit()
            flash('人员信息创建成功！', 'success')
            return redirect(url_for('personnel.personnel_list'))
        except Exception as e:
            flash(f'创建人员信息失败: {str(e)}', 'danger')
    return render_template('personnel/personnel_form.html', form=form, legend='添加人员')

@bp.route('/personnel/<int:person_id>/update', methods=['GET', 'POST'])
@department_permission_required
def update_personnel(person_id):
    person = Personnel.query.get_or_404(person_id)
    form = PersonnelForm()
    if form.validate_on_submit():
        try:
            # 部门信息
            person.company_dept = form.company_dept.data
            person.factory_dept = form.factory_dept.data
            person.dept_code = form.dept_code.data
            person.dept_full_name = form.dept_full_name.data
            person.business_unit = form.business_unit.data
            person.first_level_dept_abbr = form.first_level_dept_abbr.data
            person.org_category = form.org_category.data
            person.job_family = form.job_family.data
            person.dept_level1 = form.dept_level1.data
            person.dept_level2 = form.dept_level2.data
            person.dept_level3 = form.dept_level3.data
            person.dept_level4 = form.dept_level4.data
            person.leaf_dept = form.leaf_dept.data
            person.region_org = form.region_org.data
            person.work_location = form.work_location.data
            
            # 基本信息
            person.emp_id = form.emp_id.data
            person.emp_name = form.emp_name.data
            person.emp_status = form.emp_status.data
            person.emp_gender = form.emp_gender.data
            person.emp_type = form.emp_type.data
            person.group_hire_date = form.group_hire_date.data
            person.hire_date = form.hire_date.data
            person.is_probation = form.is_probation.data
            person.probation_end_date = form.probation_end_date.data
            person.regular_date = form.regular_date.data
            
            # 职位信息
            person.position = form.position.data
            person.position_category = form.position_category.data
            person.position_grade = form.position_grade.data
            person.job_type = form.job_type.data
            person.job_category = form.job_category.data
            person.job_sequence = form.job_sequence.data
            person.job_title = form.job_title.data
            person.acting_appointment = form.acting_appointment.data
            person.project_job = form.project_job.data
            person.job_rank = form.job_rank.data
            person.daily_salary_grade = form.daily_salary_grade.data
            person.qualification_grade = form.qualification_grade.data
            person.professional_grade = form.professional_grade.data
            person.job_level = form.job_level.data
            person.project_job_level = form.project_job_level.data
            person.talent_tag = form.talent_tag.data
            person.is_mentor = form.is_mentor.data
            
            # 汇报关系
            person.report_superior_id = form.report_superior_id.data
            person.report_superior = form.report_superior.data
            person.dept_head_id = form.dept_head_id.data
            person.dept_head = form.dept_head.data
            
            # 个人信息
            person.nationality = form.nationality.data
            person.birthplace = form.birthplace.data
            person.ethnicity = form.ethnicity.data
            person.marital_status = form.marital_status.data
            person.political_status = form.political_status.data
            
            # 教育信息
            person.highest_education_type = form.highest_education_type.data
            person.highest_education_mode = form.highest_education_mode.data
            person.highest_education = form.highest_education.data
            person.school = form.school.data
            person.major = form.major.data
            person.education_start_date = form.education_start_date.data
            person.education_end_date = form.education_end_date.data
            person.first_education = form.first_education.data
            person.second_education = form.second_education.data
            
            # 身份信息
            person.birth_date = form.birth_date.data
            person.age = form.age.data
            person.id_type = form.id_type.data
            person.id_number = form.id_number.data
            person.id_address = form.id_address.data
            person.id_expiry = form.id_expiry.data
            
            # 合同信息
            person.legal_entity = form.legal_entity.data
            person.contract_entity = form.contract_entity.data
            person.contract_type = form.contract_type.data
            person.contract_start_date = form.contract_start_date.data
            person.contract_end_date = form.contract_end_date.data
            person.labor_company = form.labor_company.data
            
            # 联系信息
            person.current_residence = form.current_residence.data
            person.registered_residence = form.registered_residence.data
            person.phone_number = form.phone_number.data
            person.emergency_contact_name = form.emergency_contact_name.data
            person.emergency_contact_phone = form.emergency_contact_phone.data
            
            # 财务信息
            person.bank_name = form.bank_name.data
            person.bank_branch = form.bank_branch.data
            person.bank_account = form.bank_account.data
            person.financial_code = form.financial_code.data
            
            # 其他信息
            person.union_member = form.union_member.data
            person.is_veteran = form.is_veteran.data
            person.initial_social_insurance_unit = form.initial_social_insurance_unit.data
            person.has_fuel_card = form.has_fuel_card.data
            
            # 离职信息
            person.resignation_date = form.resignation_date.data
            person.last_pay_date = form.last_pay_date.data
            person.resignation_type = form.resignation_type.data
            person.resignation_reason = form.resignation_reason.data
            person.is_blacklisted = form.is_blacklisted.data
            person.resignation_reason_detail = form.resignation_reason_detail.data
            person.resignation_interview = form.resignation_interview.data
            
            db.session.commit()
            flash('人员信息更新成功！', 'success')
            return redirect(url_for('personnel.personnel_list'))
        except Exception as e:
            flash(f'更新人员信息失败: {str(e)}', 'danger')
    elif request.method == 'GET':
        # 部门信息
        form.company_dept.data = person.company_dept
        form.factory_dept.data = person.factory_dept
        form.dept_code.data = person.dept_code
        form.dept_full_name.data = person.dept_full_name
        form.business_unit.data = person.business_unit
        form.first_level_dept_abbr.data = person.first_level_dept_abbr
        form.org_category.data = person.org_category
        form.job_family.data = person.job_family
        form.dept_level1.data = person.dept_level1
        form.dept_level2.data = person.dept_level2
        form.dept_level3.data = person.dept_level3
        form.dept_level4.data = person.dept_level4
        form.leaf_dept.data = person.leaf_dept
        form.region_org.data = person.region_org
        form.work_location.data = person.work_location
        
        # 基本信息
        form.emp_id.data = person.emp_id
        form.emp_name.data = person.emp_name
        form.emp_status.data = person.emp_status
        form.emp_gender.data = person.emp_gender
        form.emp_type.data = person.emp_type
        form.group_hire_date.data = person.group_hire_date
        form.hire_date.data = person.hire_date
        form.is_probation.data = person.is_probation
        form.probation_end_date.data = person.probation_end_date
        form.regular_date.data = person.regular_date
        
        # 职位信息
        form.position.data = person.position
        form.position_category.data = person.position_category
        form.position_grade.data = person.position_grade
        form.job_type.data = person.job_type
        form.job_category.data = person.job_category
        form.job_sequence.data = person.job_sequence
        form.job_title.data = person.job_title
        form.acting_appointment.data = person.acting_appointment
        form.project_job.data = person.project_job
        form.job_rank.data = person.job_rank
        form.daily_salary_grade.data = person.daily_salary_grade
        form.qualification_grade.data = person.qualification_grade
        form.professional_grade.data = person.professional_grade
        form.job_level.data = person.job_level
        form.project_job_level.data = person.project_job_level
        form.talent_tag.data = person.talent_tag
        form.is_mentor.data = person.is_mentor
        
        # 汇报关系
        form.report_superior_id.data = person.report_superior_id
        form.report_superior.data = person.report_superior
        form.dept_head_id.data = person.dept_head_id
        form.dept_head.data = person.dept_head
        
        # 个人信息
        form.nationality.data = person.nationality
        form.birthplace.data = person.birthplace
        form.ethnicity.data = person.ethnicity
        form.marital_status.data = person.marital_status
        form.political_status.data = person.political_status
        
        # 教育信息
        form.highest_education_type.data = person.highest_education_type
        form.highest_education_mode.data = person.highest_education_mode
        form.highest_education.data = person.highest_education
        form.school.data = person.school
        form.major.data = person.major
        form.education_start_date.data = person.education_start_date
        form.education_end_date.data = person.education_end_date
        form.first_education.data = person.first_education
        form.second_education.data = person.second_education
        
        # 身份信息
        form.birth_date.data = person.birth_date
        form.age.data = person.age
        form.id_type.data = person.id_type
        form.id_number.data = person.id_number
        form.id_address.data = person.id_address
        form.id_expiry.data = person.id_expiry
        
        # 合同信息
        form.legal_entity.data = person.legal_entity
        form.contract_entity.data = person.contract_entity
        form.contract_type.data = person.contract_type
        form.contract_start_date.data = person.contract_start_date
        form.contract_end_date.data = person.contract_end_date
        form.labor_company.data = person.labor_company
        
        # 联系信息
        form.current_residence.data = person.current_residence
        form.registered_residence.data = person.registered_residence
        form.phone_number.data = person.phone_number
        form.emergency_contact_name.data = person.emergency_contact_name
        form.emergency_contact_phone.data = person.emergency_contact_phone
        
        # 财务信息
        form.bank_name.data = person.bank_name
        form.bank_branch.data = person.bank_branch
        form.bank_account.data = person.bank_account
        form.financial_code.data = person.financial_code
        
        # 其他信息
        form.union_member.data = person.union_member
        form.is_veteran.data = person.is_veteran
        form.initial_social_insurance_unit.data = person.initial_social_insurance_unit
        form.has_fuel_card.data = person.has_fuel_card
        
        # 离职信息
        form.resignation_date.data = person.resignation_date
        form.last_pay_date.data = person.last_pay_date
        form.resignation_type.data = person.resignation_type
        form.resignation_reason.data = person.resignation_reason
        form.is_blacklisted.data = person.is_blacklisted
        form.resignation_reason_detail.data = person.resignation_reason_detail
        form.resignation_interview.data = person.resignation_interview
    return render_template('personnel/personnel_form.html', form=form, legend='更新人员')

@bp.route('/personnel/<int:person_id>')
@department_permission_required
def view_personnel(person_id):
    person = Personnel.query.get_or_404(person_id)
    return render_template('personnel/personnel_detail.html', person=person)

@bp.route('/personnel/<int:person_id>/delete', methods=['POST'])
@department_permission_required
def delete_personnel(person_id):
    person = Personnel.query.get_or_404(person_id)
    try:
        # 记录被删除人员信息用于日志
        emp_name = person.emp_name
        emp_id = person.emp_id
        
        # 级联清理相关数据：保密人员记录
        from app.models import ClassifiedPersonnel
        classified = ClassifiedPersonnel.query.filter_by(emp_id=emp_id).first()
        if classified:
            db.session.delete(classified)
        
        # 删除人员主体
        db.session.delete(person)
        db.session.commit()
        flash(f'人员 "{emp_name}"（工号: {emp_id}）已成功删除！', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'删除失败: {str(e)}', 'danger')
    return redirect(url_for('personnel.personnel_list'))


# ==================== 人员导入导出功能 ====================

# 常用字段定义（用于导入导出）
PERSONNEL_EXPORT_FIELDS = [
    ('工号', 'emp_id'),
    ('姓名', 'emp_name'),
    ('状态', 'emp_status'),
    ('性别', 'emp_gender'),
    ('部门代码', 'dept_code'),
    ('部门全称', 'dept_full_name'),
    ('二级部门', 'dept_level2'),
    ('职位', 'position'),
    ('职位类别', 'position_category'),
    ('职级', 'position_grade'),
    ('入职日期', 'hire_date'),
    ('手机号码', 'phone_number'),
    ('紧急联系人', 'emergency_contact_name'),
    ('紧急联系人电话', 'emergency_contact_phone'),
    ('最高学历', 'highest_education'),
    ('毕业学校', 'school'),
    ('专业', 'major'),
    ('出生日期', 'birth_date'),
    ('年龄', 'age'),
    ('身份证号', 'id_number'),
]

@bp.route('/personnel/template')
@login_required
def download_personnel_template():
    """下载人员导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人员导入模板"
    
    # 设置样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, (header_name, _) in enumerate(PERSONNEL_EXPORT_FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入示例数据
    sample_data = [
        '10001', '张三', '在职', '男', 'D001', '技术部/研发部', '研发部',
        '高级工程师', '技术', 'P7', '2020-01-15', '13800138000',
        '李四', '13900139000', '本科', '清华大学', '计算机科学',
        '1990-05-20', '34', '110101199005201234'
    ]
    
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
    
    # 设置列宽
    column_widths = [12, 10, 8, 6, 10, 25, 12, 15, 10, 8, 12, 15, 12, 15, 10, 15, 12, 12, 6, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    
    # 添加说明sheet
    ws2 = wb.create_sheet(title="填写说明")
    instructions = [
        "【人员导入模板填写说明】",
        "",
        "1. 工号：员工唯一标识，不能重复",
        "2. 状态：在职/离职/试用期",
        "3. 部门代码：需与系统中的部门代码一致",
        "4. 日期格式：YYYY-MM-DD（如：2020-01-15）",
        "5. 手机号码：11位数字",
        "6. 数据从第2行开始填写，第1行为表头请勿修改",
        "",
        "注意事项：",
        "- 导入时会根据工号判断是否已存在",
        "- 已存在的工号会跳过，不会覆盖原有数据",
        "- 请勿删除或调整列的顺序"
    ]
    
    for row, text in enumerate(instructions, 1):
        ws2.cell(row=row, column=1, value=text)
    
    ws2.column_dimensions['A'].width = 50
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'人员导入模板_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@bp.route('/personnel/export')
@login_required
def export_personnel():
    """导出人员数据"""
    search = request.args.get('search', '')
    
    query = Personnel.query
    
    # 搜索过滤
    if search:
        query = query.filter(
            Personnel.emp_name.contains(search) |
            Personnel.emp_id.contains(search) |
            Personnel.dept_full_name.contains(search) |
            Personnel.position.contains(search) |
            Personnel.phone_number.contains(search)
        )
    
    # 根据部门权限过滤
    if current_user.department_access:
        from app.decorators import get_user_accessible_departments, get_department_codes_recursive
        accessible_departments = get_user_accessible_departments()
        if accessible_departments:
            department_codes = []
            for dept_id in accessible_departments:
                department_codes.extend(get_department_codes_recursive(dept_id))
            query = query.filter(Personnel.dept_code.in_(department_codes))
    
    personnel = query.order_by(Personnel.id.desc()).all()
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人员数据"
    
    # 设置样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, (header_name, _) in enumerate(PERSONNEL_EXPORT_FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row, person in enumerate(personnel, 2):
        for col, (_, field) in enumerate(PERSONNEL_EXPORT_FIELDS, 1):
            value = getattr(person, field, None)
            cell = ws.cell(row=row, column=col, value=value if value else '')
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    
    # 设置列宽
    column_widths = [12, 10, 8, 6, 10, 25, 12, 15, 10, 8, 12, 15, 12, 15, 10, 15, 12, 12, 6, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'人员数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@bp.route('/personnel/import', methods=['POST'])
@login_required
def import_personnel():
    """导入人员数据"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '只支持 .xlsx 或 .xls 格式的文件'}), 400

    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    max_size = 10 * 1024 * 1024
    if file_size > max_size:
        return jsonify({'success': False, 'message': f'文件过大，最大支持{max_size // 1024 // 1024}MB'}), 400
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # 字段映射（中文 -> 数据库字段）
        field_mapping = {name: field for name, field in PERSONNEL_EXPORT_FIELDS}
        
        # 读取表头
        headers = {}
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if header_value and header_value in field_mapping:
                headers[field_mapping[header_value]] = col
        
        # 检查必填字段
        if 'emp_id' not in headers:
            return jsonify({'success': False, 'message': '模板格式错误：缺少"工号"列'}), 400
        
        success_count = 0
        skip_count = 0
        error_count = 0
        errors = []
        
        for row in range(2, ws.max_row + 1):
            try:
                emp_id = ws.cell(row=row, column=headers.get('emp_id', 0)).value
                
                if not emp_id:
                    continue
                
                emp_id = str(emp_id).strip()
                
                # 检查是否已存在
                existing = Personnel.query.filter_by(emp_id=emp_id).first()
                if existing:
                    skip_count += 1
                    continue
                
                # 创建新记录
                person = Personnel()
                person.emp_id = emp_id
                
                # 设置其他字段
                for field, col in headers.items():
                    if field != 'emp_id' and col > 0:
                        value = ws.cell(row=row, column=col).value
                        if value is not None:
                            setattr(person, field, str(value).strip() if isinstance(value, str) else value)
                
                db.session.add(person)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"第{row}行: {str(e)}")
        
        db.session.commit()
        
        message = f'导入完成！成功: {success_count}条，跳过(已存在): {skip_count}条'
        if error_count > 0:
            message += f'，失败: {error_count}条'
        
        return jsonify({
            'success': True,
            'message': message,
            'success_count': success_count,
            'skip_count': skip_count,
            'error_count': error_count,
            'errors': errors[:10]
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'}), 500
