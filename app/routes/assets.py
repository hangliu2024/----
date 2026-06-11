from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from datetime import datetime
import os
import io
import time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from app import db
from app.models import Personnel, ComputerInfo, IndustrialComputer
from app.decorators import department_permission_required, department_data_filter

bp = Blueprint('assets', __name__)

# 简单内存缓存 - 300秒(5分钟)过期
_dashboard_cache = {'data': None, 'timestamp': 0}
CACHE_TTL = 300

# 允许的文件扩展名
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@bp.route('/')
@bp.route('/dashboard')
@login_required
def dashboard():
    now = time.time()
    if _dashboard_cache['data'] is not None and (now - _dashboard_cache['timestamp']) < CACHE_TTL:
        cached = _dashboard_cache['data']
        return render_template('dashboard.html', **cached)

    from app.models import Personnel, ComputerInfo, IndustrialComputer, Department, LoginLog
    
    office_computers_count = ComputerInfo.query.count()
    industrial_computers_count = IndustrialComputer.query.count()
    personnel_count = Personnel.query.filter_by(emp_status='在职').count()
    personnel_total = Personnel.query.count()
    dept_count = Department.query.count()
    login_logs = LoginLog.query.order_by(LoginLog.login_time.desc()).limit(10).all()
    
    # 操作系统分布概览
    os_counts = {}
    for c in ComputerInfo.query.with_entities(ComputerInfo.operating_system).all():
        os = c.operating_system[:10] if c.operating_system else '未知'
        os_counts[os] = os_counts.get(os, 0) + 1
    os_summary = sorted(os_counts.items(), key=lambda x: -x[1])[:5]

    data = {
        'office_computers_count': office_computers_count,
        'industrial_computers_count': industrial_computers_count,
        'personnel_count': personnel_count,
        'personnel_total': personnel_total,
        'dept_count': dept_count,
        'login_logs': login_logs,
        'os_summary': os_summary
    }
    _dashboard_cache['data'] = data
    _dashboard_cache['timestamp'] = now

    return render_template('dashboard.html', **data)

# 涉密资产管理 - 办公电脑
@bp.route('/office_computers')
@department_permission_required
def office_computers():
    search = request.args.get('search', '').strip()[:100]
    page = max(request.args.get('page', 1, type=int), 1)
    per_page = max(min(request.args.get('per_page', 50, type=int), 200), 10)
    
    query = ComputerInfo.query.order_by(ComputerInfo.id.desc())
    
    # 根据部门权限过滤
    if current_user.department_access:
        from app.models import Department
        user_department = Department.query.get(current_user.department_id)
        if user_department:
            query = query.filter(ComputerInfo.dept_code == user_department.code)
    
    # 搜索过滤
    if search:
        search_filter = (
            ComputerInfo.computer_name.ilike(f'%{search}%') |
            ComputerInfo.network_address.ilike(f'%{search}%') |
            ComputerInfo.ip_mac.ilike(f'%{search}%') |
            ComputerInfo.operating_system.ilike(f'%{search}%') |
            ComputerInfo.last_login_user.ilike(f'%{search}%') |
            ComputerInfo.dept_code.ilike(f'%{search}%') |
            ComputerInfo.emp_name.ilike(f'%{search}%')
        )
        query = query.filter(search_filter)
    
    # 先获取总数
    total = query.count()
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    items = pagination.items
    
    return render_template('assets/office_computers.html', computer_infos=items, pagination=pagination, search=search, per_page=per_page, total=total)

# 工控机管理
@bp.route('/industrial_computers')
@department_permission_required
def industrial_computers():
    search = request.args.get('search', '').strip()[:100] if request.args.get('search') else None
    page = max(request.args.get('page', 1, type=int), 1)
    per_page = 50
    
    query = IndustrialComputer.query.order_by(IndustrialComputer.id.desc())
    
    if search:
        search_filter = (
            IndustrialComputer.device_name.ilike(f'%{search}%') |
            IndustrialComputer.ip_address.ilike(f'%{search}%') |
            IndustrialComputer.bu_dept.ilike(f'%{search}%') |
            IndustrialComputer.factory.ilike(f'%{search}%') |
            IndustrialComputer.location.ilike(f'%{search}%')
        )
        query = query.filter(search_filter)
    
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    items = pagination.items
    
    return render_template('assets/industrial_computers.html', items=items, pagination=pagination, search=search)


# ==================== 办公电脑导入导出功能 ====================

@bp.route('/office_computers/template')
@login_required
def download_computer_template():
    """下载办公电脑导入模板"""
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "办公电脑导入模板"
    
    # 定义表头（中文名 -> 数据库字段映射）
    headers = [
        ('电脑名称*', 'computer_name'),
        ('工号', 'employee_id'),
        ('部门代码', 'dept_code'),
        ('二级部门', 'dept_level2'),
        ('员工姓名', 'emp_name'),
        ('资产ID', 'asset_id'),
        ('网络地址', 'network_address'),
        ('IP/MAC', 'ip_mac'),
        ('操作系统', 'operating_system'),
        ('最后登录用户', 'last_login_user')
    ]
    
    # 设置样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, (header_name, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入示例数据（第2行）
    sample_data = [
        'PC-001',      # 电脑名称
        '10001',       # 工号
        'D001',        # 部门代码
        '技术部',       # 二级部门
        '张三',        # 员工姓名
        '12345',       # 资产ID
        '192.168.1.100',  # 网络地址
        '192.168.1.100 / AA:BB:CC:DD:EE:FF',  # IP/MAC
        'Windows 10 专业版',  # 操作系统
        'zhangsan'     # 最后登录用户
    ]
    
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
    
    # 设置列宽
    column_widths = [15, 12, 12, 15, 12, 12, 18, 30, 20, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    
    # 添加说明sheet
    ws2 = wb.create_sheet(title="填写说明")
    instructions = [
        "【办公电脑导入模板填写说明】",
        "",
        "1. 必填字段：电脑名称（标*号的字段）",
        "2. 工号：员工工号，用于关联人员信息",
        "3. 部门代码：部门编码，需与系统中的部门代码一致",
        "4. IP/MAC：格式为 'IP地址 / MAC地址'，例如：192.168.1.100 / AA:BB:CC:DD:EE:FF",
        "5. 数据从第2行开始填写，第1行为表头请勿修改",
        "6. 请勿删除或调整列的顺序",
        "",
        "注意事项：",
        "- 电脑名称不能重复",
        "- 导入时会自动跳过已存在的电脑名称",
        "- 日期格式：YYYY-MM-DD"
    ]
    
    for row, text in enumerate(instructions, 1):
        ws2.cell(row=row, column=1, value=text)
    
    ws2.column_dimensions['A'].width = 60
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'办公电脑导入模板_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@bp.route('/office_computers/export')
@login_required
def export_computers():
    """导出办公电脑数据"""
    # 获取查询参数
    search = request.args.get('search', '')
    
    query = ComputerInfo.query.order_by(ComputerInfo.id.desc())
    
    # 根据部门权限过滤
    if current_user.department_access:
        from app.models import Department
        user_department = Department.query.get(current_user.department_id)
        if user_department:
            query = query.filter(ComputerInfo.dept_code == user_department.code)
    
    # 搜索过滤
    if search:
        search_filter = (
            ComputerInfo.computer_name.ilike(f'%{search}%') |
            ComputerInfo.network_address.ilike(f'%{search}%') |
            ComputerInfo.ip_mac.ilike(f'%{search}%') |
            ComputerInfo.operating_system.ilike(f'%{search}%') |
            ComputerInfo.last_login_user.ilike(f'%{search}%') |
            ComputerInfo.dept_code.ilike(f'%{search}%') |
            ComputerInfo.emp_name.ilike(f'%{search}%')
        )
        query = query.filter(search_filter)
    
    # 限制导出数量，防止内存溢出
    MAX_EXPORT_ROWS = 10000
    total_count = query.count()
    computers = query.limit(MAX_EXPORT_ROWS).all()
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "办公电脑数据"
    
    # 定义表头
    headers = [
        ('ID', 'id'),
        ('电脑名称', 'computer_name'),
        ('工号', 'employee_id'),
        ('部门代码', 'dept_code'),
        ('二级部门', 'dept_level2'),
        ('员工姓名', 'emp_name'),
        ('资产ID', 'asset_id'),
        ('网络地址', 'network_address'),
        ('IP/MAC', 'ip_mac'),
        ('操作系统', 'operating_system'),
        ('最后登录用户', 'last_login_user')
    ]
    
    # 设置样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, (header_name, _) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row, computer in enumerate(computers, 2):
        data = [
            computer.id,
            computer.computer_name or '',
            computer.employee_id or '',
            computer.dept_code or '',
            computer.dept_level2 or '',
            computer.emp_name or '',
            computer.asset_id or '',
            computer.network_address or '',
            computer.ip_mac or '',
            computer.operating_system or '',
            computer.last_login_user or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    
    # 设置列宽
    column_widths = [8, 18, 12, 12, 15, 12, 12, 18, 30, 20, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    download_name = f'办公电脑数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    if total_count > MAX_EXPORT_ROWS:
        download_name = f'办公电脑数据_前{MAX_EXPORT_ROWS}条_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )


@bp.route('/office_computers/import', methods=['POST'])
@login_required
@department_permission_required
def import_computers():
    """导入办公电脑数据"""
    # 检查是否有文件
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'}), 400
    
    file = request.files['file']
    
    # 检查文件名
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'}), 400
    
    # 检查文件类型
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '只支持 .xlsx 或 .xls 格式的文件'}), 400
    
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    max_size = 10 * 1024 * 1024
    if file_size > max_size:
        return jsonify({'success': False, 'message': f'文件过大，最大支持{max_size // 1024 // 1024}MB'}), 400
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        field_mapping = {
            '电脑名称*': 'computer_name',
            '电脑名称': 'computer_name',
            '工号': 'employee_id',
            '部门代码': 'dept_code',
            '二级部门': 'dept_level2',
            '员工姓名': 'emp_name',
            '资产ID': 'asset_id',
            '网络地址': 'network_address',
            'IP/MAC': 'ip_mac',
            '操作系统': 'operating_system',
            '最后登录用户': 'last_login_user'
        }
        
        # 读取表头
        headers = {}
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if header_value and header_value in field_mapping:
                headers[field_mapping[header_value]] = col
        
        # 检查必填字段
        if 'computer_name' not in headers:
            return jsonify({'success': False, 'message': '模板格式错误：缺少"电脑名称"列'}), 400
        
        # 导入数据
        success_count = 0
        skip_count = 0
        error_count = 0
        errors = []
        
        for row in range(2, ws.max_row + 1):
            try:
                # 读取行数据
                computer_name = ws.cell(row=row, column=headers.get('computer_name', 0)).value
                
                # 跳过空行
                if not computer_name:
                    continue
                
                # 检查是否已存在
                existing = ComputerInfo.query.filter_by(computer_name=str(computer_name).strip()).first()
                if existing:
                    skip_count += 1
                    continue
                
                # 创建新记录
                computer = ComputerInfo()
                computer.computer_name = str(computer_name).strip() if computer_name else None
                
                # 设置其他字段
                for field, col in headers.items():
                    if field != 'computer_name' and col > 0:
                        value = ws.cell(row=row, column=col).value
                        if value is not None:
                            setattr(computer, field, str(value).strip() if isinstance(value, str) else value)
                
                db.session.add(computer)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"第{row}行: {str(e)}")
        
        # 提交事务
        db.session.commit()
        
        message = f'导入完成！成功: {success_count}条，跳过(已存在): {skip_count}条'
        if error_count > 0:
            message += f'，失败: {error_count}条'
        
        return jsonify({
            'success': True,
            'message': message,
            'success_count': success_count,
            'skip_count': skip_count,
            'error_count': error_count,
            'errors': errors[:10]  # 只返回前10条错误
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'}), 500


# ==================== 工控机导入导出功能 ====================

# 工控机导出字段定义
INDUSTRIAL_COMPUTER_FIELDS = [
    ('序号', 'seq_no'),
    ('区域', 'zone'),
    ('BU部门', 'bu_dept'),
    ('工厂', 'factory'),
    ('设备名称', 'device_name'),
    ('位置/工序', 'location'),
    ('操作系统', 'operating_system'),
    ('IP地址', 'ip_address'),
    ('MAC地址', 'mac_address'),
    ('第三方杀毒', 'third_party_antivirus'),
    ('IEP安装', 'iep_installed'),
]

@bp.route('/industrial_computers/template')
@login_required
def download_industrial_template():
    """下载工控机导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工控机导入模板"
    
    # 设置样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, (header_name, _) in enumerate(INDUSTRIAL_COMPUTER_FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入示例数据
    sample_data = [
        1, '华东', '制造部', '上海工厂', 'IPC-001', '包装线A区',
        'Windows 10 IoT', '192.168.1.100', 'AA:BB:CC:DD:EE:FF', '是', '是'
    ]
    
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
    
    # 设置列宽
    column_widths = [8, 10, 15, 12, 15, 15, 18, 15, 20, 12, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    
    # 添加说明sheet
    ws2 = wb.create_sheet(title="填写说明")
    instructions = [
        "【工控机导入模板填写说明】",
        "",
        "1. 设备名称：工控机唯一标识，不能重复",
        "2. 区域：设备所在区域",
        "3. 第三方杀毒：是/否",
        "4. IEP安装：是/否",
        "5. IP地址：格式如 192.168.1.100",
        "6. MAC地址：格式如 AA:BB:CC:DD:EE:FF",
        "7. 数据从第2行开始填写，第1行为表头请勿修改",
        "",
        "注意事项：",
        "- 导入时会根据设备名称判断是否已存在",
        "- 已存在的设备名称会跳过，不会覆盖原有数据",
        "- 请勿删除或调整列的顺序"
    ]
    
    for row, text in enumerate(instructions, 1):
        ws2.cell(row=row, column=1, value=text)
    
    ws2.column_dimensions['A'].width = 50
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'工控机导入模板_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@bp.route('/industrial_computers/export')
@login_required
def export_industrial_computers():
    """导出工控机数据"""
    search = request.args.get('search', '')
    
    query = IndustrialComputer.query
    
    if search:
        search_filter = (
            IndustrialComputer.device_name.ilike(f'%{search}%') |
            IndustrialComputer.ip_address.ilike(f'%{search}%') |
            IndustrialComputer.bu_dept.ilike(f'%{search}%') |
            IndustrialComputer.factory.ilike(f'%{search}%') |
            IndustrialComputer.location.ilike(f'%{search}%')
        )
        query = query.filter(search_filter)
    
    # 限制导出数量，防止内存溢出
    MAX_EXPORT_ROWS = 10000
    total_count = query.count()
    items = query.order_by(IndustrialComputer.id.desc()).limit(MAX_EXPORT_ROWS).all()
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工控机数据"
    
    # 设置样式
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, (header_name, _) in enumerate(INDUSTRIAL_COMPUTER_FIELDS, 1):
        cell = ws.cell(row=1, column=col, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 写入数据
    for row, item in enumerate(items, 2):
        for col, (_, field) in enumerate(INDUSTRIAL_COMPUTER_FIELDS, 1):
            value = getattr(item, field, None)
            cell = ws.cell(row=row, column=col, value=value if value else '')
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    
    # 设置列宽
    column_widths = [8, 10, 15, 12, 15, 15, 18, 15, 20, 12, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    download_name = f'工控机数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    if total_count > MAX_EXPORT_ROWS:
        download_name = f'工控机数据_前{MAX_EXPORT_ROWS}条_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=download_name
    )


@bp.route('/industrial_computers/import', methods=['POST'])
@login_required
@department_permission_required
def import_industrial_computers():
    """导入工控机数据"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有选择文件'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'message': '只支持 .xlsx 或 .xls 格式的文件'}), 400
    
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    if file_size > 10 * 1024 * 1024:
        return jsonify({'success': False, 'message': '文件过大，最大支持10MB'}), 400
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        field_mapping = {name: field for name, field in INDUSTRIAL_COMPUTER_FIELDS}
        
        # 读取表头
        headers = {}
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if header_value and header_value in field_mapping:
                headers[field_mapping[header_value]] = col
        
        # 检查必填字段（设备名称）
        if 'device_name' not in headers:
            return jsonify({'success': False, 'message': '模板格式错误：缺少"设备名称"列'}), 400
        
        success_count = 0
        skip_count = 0
        error_count = 0
        errors = []
        
        for row in range(2, ws.max_row + 1):
            try:
                device_name = ws.cell(row=row, column=headers.get('device_name', 0)).value
                
                if not device_name:
                    continue
                
                device_name = str(device_name).strip()
                
                # 检查是否已存在
                existing = IndustrialComputer.query.filter_by(device_name=device_name).first()
                if existing:
                    skip_count += 1
                    continue
                
                # 创建新记录
                item = IndustrialComputer()
                item.device_name = device_name
                
                # 设置其他字段
                for field, col in headers.items():
                    if field != 'device_name' and col > 0:
                        value = ws.cell(row=row, column=col).value
                        if value is not None:
                            setattr(item, field, str(value).strip() if isinstance(value, str) else value)
                
                db.session.add(item)
                success_count += 1
                
            except Exception as e:
                error_count += 1
                errors.append(f"第{row}行: {str(e)}")
        
        db.session.commit()
        
        message = f'导入完成！成功: {success_count}条，跳过(已存在): {skip_count}条'
        if error_count > 0:
            message += f'，失败: {error_count}条'
        
        return jsonify({
            'success': True,
            'message': message,
            'success_count': success_count,
            'skip_count': skip_count,
            'error_count': error_count,
            'errors': errors[:10]
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'}), 500
