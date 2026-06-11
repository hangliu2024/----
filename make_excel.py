"""生成专业软件开发项目管理表"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ============================================================
# Sheet 1: 项目总览
# ============================================================
ws = wb.active
ws.title = "项目总览"
ws.sheet_properties.tabColor = "1a365d"

# 颜色定义
NAVY = "1a365d"
BLUE = "0052aa"
ORANGE = "e66400"
LIGHT = "f0f4ff"
WHITE = "ffffff"
GRAY = "f5f5f5"
DARK = "333333"

hdr_font = Font(name="微软雅黑", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
title_font = Font(name="微软雅黑", bold=True, color=NAVY, size=16)
sub_font = Font(name="微软雅黑", color=DARK, size=10)
data_font = Font(name="微软雅黑", color=DARK, size=10)
link_font = Font(name="微软雅黑", color=ORANGE, size=10, bold=True)
thin_border = Border(
    left=Side(style="thin", color="d0d0d0"),
    right=Side(style="thin", color="d0d0d0"),
    top=Side(style="thin", color="d0d0d0"),
    bottom=Side(style="thin", color="d0d0d0"),
)

# 标题行
ws.merge_cells("A1:H1")
ws["A1"] = "企业资产与信息安全综合管理平台 — 项目开发管理表"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:H2")
ws["A2"] = f"项目经理：XXX  |  项目周期：2026年1月 - 2026年6月  |  状态：收尾阶段  |  更新日期：2026年6月3日"
ws["A2"].font = sub_font
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 24

# 概要统计行
stats = [
    ("总任务数", "28", "已完成", "18", "进行中", "6", "未开始", "4"),
]
for i, v in enumerate(stats[0]):
    col = get_column_letter(i + 1)
    cell = ws[f"A4"]
ws.merge_cells("A4:H4")

# 用一行显示进度
ws["A4"] = "总任务：28项  |  已完成：18项 (64%)  |  进行中：6项 (21%)  |  未开始：4项 (14%)  |  预计完成：2026年6月7日"
ws["A4"].font = Font(name="微软雅黑", bold=True, color=ORANGE, size=11)
ws["A4"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[4].height = 28

# 表头
headers = ["编号", "任务名称", "负责人", "优先级", "状态", "进度", "计划日期", "备注"]
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=6, column=i, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[6].height = 32

# 任务数据（根据实际项目真实填写）
tasks = [
    # (编号, 名称, 负责人, 优先级, 状态, 进度, 日期, 备注)
    ("P1", "需求分析与系统设计", "产品/架构", "高", "已完成", "100%", "1月-2月", "完成PRD和技术方案"),
    ("P2", "数据库设计与建表", "后端组", "高", "已完成", "100%", "2月", "11张核心业务表"),
    ("P3", "用户认证模块（登录/注册/权限）", "后端组", "高", "已完成", "100%", "2月-3月", "JWT+CSRF+7级RBAC"),
    ("P4", "人员管理模块", "后端组", "高", "已完成", "100%", "3月", "93K员工档案管理"),
    ("P5", "IT资产管理（办公电脑）", "后端组", "高", "已完成", "100%", "3月", "1000+台设备纳管"),
    ("P6", "IT资产管理（工控机）", "后端组", "高", "已完成", "100%", "3月", "500+台工控机管理"),
    ("P7", "涉密安全管理模块", "后端组", "高", "已完成", "100%", "3月-4月", "人/物/区域/文件管控"),
    ("P8", "AI智能问答模块", "AI组", "高", "已完成", "100%", "4月", "自然语言查数据"),
    ("P9", "权限安全加固", "安全组", "高", "已完成", "100%", "4月", "密码策略+CSRF+XSS修复"),
    ("P10", "界面统一优化（小米风格）", "前端组", "中", "已完成", "100%", "4月", "全站设计统一"),
    ("P11", "稽查管理模块", "后端组", "高", "已完成", "100%", "4月", "任务分配+跟踪+记录"),
    ("P12", "案件管理模块", "后端组", "中", "已完成", "100%", "4月", "案例集+SOP+调查报告"),
    ("P13", "应急管理模块", "后端组", "中", "已完成", "100%", "4月", "预案+演练+应急小组"),
    ("P14", "部门管理与数据隔离", "后端组", "高", "已完成", "100%", "4月", "部门数据和权限隔离"),
    ("P15", "操作日志审计", "后端组", "高", "已完成", "100%", "5月", "操作留痕+登录日志"),
    ("P16", "登录日志记录", "后端组", "中", "已完成", "100%", "5月", "登录/登出记录"),
    ("P17", "仪表盘增强", "前端组", "中", "已完成", "100%", "5月", "KPI+图表+动态"),
    ("P18", "AI流式输出与对话记忆", "AI组", "高", "已完成", "100%", "5月-6月", "SSE流式+数据库存储"),
    ("P19", "系统安全测试", "安全组", "高", "进行中", "90%", "5月-6月", "代码审计+渗透测试"),
    ("P20", "全面回归测试", "测试组", "高", "进行中", "75%", "5月-6月", "功能+性能+安全验证"),
    ("P21", "数据库性能优化", "后端组", "中", "进行中", "60%", "5月-6月", "索引+字段扩容+导入效率"),
    ("P22", "文档完善（用户手册+部署文档）", "全员", "中", "进行中", "50%", "5月-6月", "操作手册+运维手册"),
    ("P23", "UX用户体验优化", "前端组", "中", "进行中", "80%", "6月", "加载状态/分页/删除确认等"),
    ("P24", "单点登录限制移除", "后端组", "低", "已完成", "100%", "6月", "允许多设备同时登录"),
    ("P25", "中文乱码修复", "后端组", "高", "已完成", "100%", "6月", "SSE流式charset修复"),
    ("P26", "AI模式切换（数据问答/对话）", "AI组", "中", "已完成", "100%", "6月", "去掉意图识别"),
    ("P27", "正式上线部署", "运维组", "高", "未开始", "0%", "6月7日", "生产环境部署"),
    ("P28", "上线后监控与运维", "运维组", "高", "未开始", "0%", "6月7日起", "7x24小时监控"),
]

# 优先级颜色映射
priority_fills = {
    "高": PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid"),
    "中": PatternFill(start_color="fff8e1", end_color="fff8e1", fill_type="solid"),
    "低": PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid"),
}
priority_fonts = {
    "高": Font(name="微软雅黑", bold=True, color="c62828", size=10),
    "中": Font(name="微软雅黑", bold=True, color="e65100", size=10),
    "低": Font(name="微软雅黑", bold=True, color="2e7d32", size=10),
}
status_fills = {
    "已完成": PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid"),
    "进行中": PatternFill(start_color="e3f2fd", end_color="e3f2fd", fill_type="solid"),
    "未开始": PatternFill(start_color="fce4ec", end_color="fce4ec", fill_type="solid"),
}
status_fonts = {
    "已完成": Font(name="微软雅黑", bold=True, color="2e7d32", size=10),
    "进行中": Font(name="微软雅黑", bold=True, color="1565c0", size=10),
    "未开始": Font(name="微软雅黑", bold=True, color="c62828", size=10),
}

for r, task in enumerate(tasks, 7):
    for c, val in enumerate(task, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if c in (1, 3, 4, 5, 6) else "left", vertical="center", wrap_text=True)
        if c == 4:  # 优先级
            cell.fill = priority_fills.get(val, PatternFill())
            cell.font = priority_fonts.get(val, data_font)
        if c == 5:  # 状态
            cell.fill = status_fills.get(val, PatternFill())
            cell.font = status_fonts.get(val, data_font)
        if c == 6 and val:  # 进度条颜色
            pct = int(val.replace("%", ""))
            if pct == 100:
                cell.fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
            elif pct >= 75:
                cell.fill = PatternFill(start_color="bbdefb", end_color="bbdefb", fill_type="solid")
            elif pct >= 50:
                cell.fill = PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid")
    # 隔行色
    if r % 2 == 0:
        for c in range(1, 9):
            if c not in (4, 5, 6):
                ws.cell(row=r, column=c).fill = PatternFill(start_color="fafafa", end_color="fafafa", fill_type="solid")
    ws.row_dimensions[r].height = 26

# 列宽
widths = [8, 38, 10, 10, 10, 10, 16, 30]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 2: 进度甘特图（简化版）
# ============================================================
ws2 = wb.create_sheet("进度甘特图")
ws2.sheet_properties.tabColor = "e66400"

ws2["A1"] = "项目进度甘特图"
ws2["A1"].font = Font(name="微软雅黑", bold=True, color=NAVY, size=16)
ws2.row_dimensions[1].height = 36

ws2.merge_cells("A3:B3")
ws2["A3"] = "任务"
ws2["A3"].font = hdr_font
ws2["A3"].fill = hdr_fill
ws2["A3"].alignment = Alignment(horizontal="center", vertical="center")
ws2["A3"].border = thin_border

# 月份列
months = ["1月", "2月", "3月", "4月", "5月", "6月"]
ws2.merge_cells("C3:H3")
ws2["C3"] = "时间"
ws2["C3"].font = hdr_font
ws2["C3"].fill = hdr_fill
ws2["C3"].alignment = Alignment(horizontal="center", vertical="center")
ws2["C3"].border = thin_border

# 子月份
for i, m in enumerate(months):
    cell = ws2.cell(row=4, column=i + 3, value=m)
    cell.font = hdr_font
    cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws2.row_dimensions[4].height = 28

# 项目时间线
timeline = [
    ("P1 需求分析", 1, 2, NAVY),
    ("P2 数据库设计", 2, 2, NAVY),
    ("P3 用户认证", 2, 2, BLUE),
    ("P4 人员管理", 3, 1, BLUE),
    ("P5-P6 资产管理", 3, 1, BLUE),
    ("P7 涉密安全", 3, 2, BLUE),
    ("P8 AI问答", 4, 1, BLUE),
    ("P9 权限加固", 4, 1, BLUE),
    ("P10-P14 其他模块", 4, 1, BLUE),
    ("P15-P18 审计/优化", 5, 1, BLUE),
    ("P19-P20 安全+回归测试", 5, 2, ORANGE),
    ("P21-P23 优化+文档", 5, 2, ORANGE),
    ("P24-P26 最后修复", 6, 0.5, ORANGE),
    ("P27-P28 上线部署", 6, 0.5, "c62828"),
]

gantt_fills = {
    NAVY: PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid"),
    BLUE: PatternFill(start_color="0052aa", end_color="0052aa", fill_type="solid"),
    ORANGE: PatternFill(start_color="e66400", end_color="e66400", fill_type="solid"),
    "c62828": PatternFill(start_color="c62828", end_color="c62828", fill_type="solid"),
}

for r, (name, start_month, duration, color) in enumerate(timeline, 5):
    cell = ws2.cell(row=r, column=1, value=name)
    cell.font = Font(name="微软雅黑", bold=True, color=DARK, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border
    ws2.merge_cells(f"A{r}:B{r}")
    if color == ORANGE:
        cell.fill = PatternFill(start_color="fff3e0", end_color="fff3e0", fill_type="solid")
    
    # 甘特条
    for d in range(int(duration * 4)):
        col = 3 + (start_month - 1) * 2 + d // 2
        if col <= 14:  # 到6月
            cell = ws2.cell(row=r, column=col, value="")
            cell.fill = gantt_fills.get(color, gantt_fills[BLUE])
            cell.border = thin_border
    ws2.row_dimensions[r].height = 22

# 图例
ws2.cell(row=20, column=1, value="图例：").font = Font(name="微软雅黑", bold=True, size=10)
legends = [("需求/设计", NAVY), ("开发", BLUE), ("测试/优化", ORANGE), ("上线", "c62828")]
for i, (name, color) in enumerate(legends):
    r = 20
    c = 2 + i * 3
    cell = ws2.cell(row=r, column=c, value="  ")
    cell.fill = gantt_fills[color]
    cell.border = thin_border
    ws2.cell(row=r, column=c + 1, value=name).font = Font(name="微软雅黑", size=10)

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 6
for i in range(3, 15):
    ws2.column_dimensions[get_column_letter(i)].width = 5

# ============================================================
# Sheet 3: 问题跟踪
# ============================================================
ws3 = wb.create_sheet("问题跟踪")
ws3.sheet_properties.tabColor = "c62828"

ws3["A1"] = "问题跟踪表"
ws3["A1"].font = Font(name="微软雅黑", bold=True, color=NAVY, size=16)
ws3.row_dimensions[1].height = 36

issues_headers = ["编号", "问题描述", "提出人", "严重程度", "状态", "提出日期", "解决人", "备注"]
for i, h in enumerate(issues_headers, 1):
    cell = ws3.cell(row=3, column=i, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws3.row_dimensions[3].height = 32

issues = [
    ("I1", "AI返回中文乱码", "测试组", "高", "已解决", "5月30日", "后端组", "charset=utf-8修复"),
    ("I2", "管理员无法同时登录", "运维组", "高", "已解决", "5月28日", "后端组", "移除单点登录限制"),
    ("I3", "办公电脑无编辑/删除", "用户", "中", "已解决", "5月25日", "后端组", "已添加"),
    ("I4", "数据库导入效率低", "运维组", "中", "进行中", "5月20日", "后端组", "批次优化"),
    ("I5", "导航栏不显示当前位置", "用户", "低", "已解决", "5月26日", "前端组", "JS高亮"),
    ("I6", "AI意图识别不准", "测试组", "中", "已解决", "6月1日", "AI组", "改为手动模式切换"),
]

fills_severity = {
    "高": PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid"),
    "中": PatternFill(start_color="fff8e1", end_color="fff8e1", fill_type="solid"),
    "低": PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid"),
}
fonts_severity = {
    "高": Font(name="微软雅黑", bold=True, color="c62828", size=10),
    "中": Font(name="微软雅黑", bold=True, color="e65100", size=10),
    "低": Font(name="微软雅黑", bold=True, color="2e7d32", size=10),
}

for r, issue in enumerate(issues, 4):
    for c, val in enumerate(issue, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if c in (1, 3, 4, 5) else "left", vertical="center", wrap_text=True)
        if c == 4:
            cell.fill = fills_severity.get(val, PatternFill())
            cell.font = fonts_severity.get(val, data_font)
        if c == 5:
            if val == "已解决":
                cell.fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
                cell.font = Font(name="微软雅黑", bold=True, color="2e7d32", size=10)
            else:
                cell.fill = PatternFill(start_color="bbdefb", end_color="bbdefb", fill_type="solid")
                cell.font = Font(name="微软雅黑", bold=True, color="1565c0", size=10)
    ws3.row_dimensions[r].height = 24

col_widths = [8, 36, 10, 10, 10, 12, 10, 28]
for i, w in enumerate(col_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# 保存
# ============================================================
path = "d:\\资产管理\\项目开发管理表.xlsx"
wb.save(path)
print(f"✅ 已生成: {path}")
